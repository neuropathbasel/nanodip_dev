#!/usr/bin/env python
# coding: utf-8

# In[1]:


versionString="53"                                   # version string of this application


# # NanoDiP all-in-one Jupyter Notebook
# *J. Hench, C. Hultschig, J. Brugger, and S. Frank, Neuropathology, IfP Basel, 2021-2023*
# 
# This software is provided free of charge and warranty; by using it you agree to do this on your own risk. The authors shall not be held liable for any damage caused by this software. We have assembled this and tested it to the best of our knowledge.
# 
# The purpose of NanoDiP (Nanopore Digital Pathology) is to compare low-coverage Nanopore sequencing data from natively extracted DNA sequencing runs against a flexibly adaptable collection of 450K/850K Illumina Infinium Methylation array data. These data have to be preprocessed into binary beta value files; this operation is performed in R (uses minfi to read raw array data) and outputs bindary float files (one per dataset). These beta values files (e.g., 204949770141_R03C01_betas_filtered.bin) are named according to the array ID (Sentrix ID) followed by the suffix. A collection of betas_filtered.bin files can be provided in a static manner and XLSX (Microsoft Excel) tables can be used to select a subset thereof alongside a user-defined annotation. The corresponding datasets will be loaded into memory and then serve as the reference cohort to which the Nanopore data are compared by dimension reduction (UMAP). This comparison is optimized for speed and low resource consumption so that it can run on the computer that operates the sequencer. The sequencing run is initiated through the MinKNOW API by this application. Basecalling and methylation calling occur as background tasks outside this Jupyter Notebook. User interaction occurs through a web interface based on CherryPy which has been tested on Chromium web browser. It is advisable to run it locally, there are no measures to secure the generated website.
# 
# In order to use this application properly please make sure to be somewhat familiar with Jupyter Notebook. To run the software, press the button called *restart the kernel, re-run the whole notebook (with dialog)* and confirm execution. Then, in Chromium Browser, navigate to http://localhost:8080/ and preferably bookmark this location for convenience. In case of errors, you may just again click the same button *restart the kernel, re-run the whole notebook (with dialog)*.
# ___
# ### Technical Details
# * Tested with Python 3.7.5; 3.8.8 fails to load minknow_api in jupyter notebook.
# * **CAUTION**: Requires a *patched* version of minknow api, file `[VENV]/lib/python3.7/site-packages/minknow_api/tools/protocols.py`. Without the patch, the generated fast5 sequencing data will be unreadable with f5c or nanopolish (wrong compression algorithm, which is the default in the MinKNOW backend).
# * Verified to run on Ubuntu 18.04/Jetpack on ARMv8 and x86_64 CPUs, both with CUDA-compatible GPU; not tested on Windows and Mac OS. The latter two platforms are unsupported, we do not intend to support them. A CUDA-compatible GPU is not required but highly recommended. The CUPY package must be installated. Installation is possible on systems without GPU. Cupy installation takes a long time (often hours). This is an initial step which involves system-specific compilation of the package. Compilation occurs in the background.
# ___
# ### Headless / Command Line Mode
# CherryPy, the underlying web server of NanoDiP facilitates headless (command line-based) operation of the software besides or instead of browser-based use. Hence, the software may be operated as a post-hoc analysis pipeline for previously acquired data. This is particularly useful for benchmarking and validation purposes.
# 
# #### Examples:
# Generate copy number of for sample **GBM_RTK2_20210311_Testrun_BC06**: \
# `curl 'http://localhost:8081/cnvplot?sampleName=GBM_RTK2_20210311_Testrun_BC06'`
# 
# Calculate UMAP plot for sample **GBM_RTK2_20210311_Testrun_BC06** with reference annotation **AllIDATv2_20210804.xlsx**: \
# `curl 'http://localhost:8081/umapplot?sampleName=GBM_RTK2_20210311_Testrun_BC06&refAnno=AllIDATv2_20210804.xlsx'`
# 
# Assemble PDF report for sample **GBM_RTK2_20210311_Testrun_BC06** with reference annotation **AllIDATv2_20210804.xlsx**: \
# `curl 'http://localhost:8081/makePdf?sampleName=GBM_RTK2_20210311_Testrun_BC06&refAnno=AllIDATv2_20210804.xlsx'`
# 

# ### Version Details
# **30:** UMAP report score / PDF (NanoDiP)
# 
# **32:** UMAP report score / PDF (EpiDiP)
# 
# **35:** EpiDiP UMAP and SD ranking on GPU, GPU/CPU system-agnostic code, runs with or w/o GPU
# 
# **37:** EpiDiP custom annotated CNV plots
# 
# **42:** Adjustment for older fast5 naming conventions from MinKNOW
# 
# **43:** Development for R9-R10 transision, switch to version 5.4.0 version of MinKNOW API
# 
# **44:** Integration of P2S (PromethION P2 solo) devices
# 
# **45:** Adaptation to MinKNOW 23.07.5 (STABLE) 2023-09-05 This requires installation of minknow-api-5.7.2. Added R9/R10 switch.
# 
# **46:** Removed basecalling during the run due to incompatibility of many guppy/dorado versions. Adaptations to also run on ARM64 SoC (ORIN AGX 32GB). Failed for python 3.8.10 (from ubuntu 20.04), switched back to 3.7 (from deadsnakes repository) due to scikit-learn not working.
# 
# **47:** Basecalling implemented in NanoDiP
# 
# **48:** Blacklisting of Illumina array IDs (for EPICv2 compatibility)
# 
# **49:** Adaptations for "minit" vs. "minknow" vs. other user's home pathes. Minor UI improvements.
# 
# **50:** Generate UMAP reports from locally generated UMAP coordinates
# 
# **51:** Attempt to exchange basecalling models and to adjust f5c methylation calling parameters
# 
# **52:** Add menu do delete basecalling and methylation calling temporary files
# 
# **53:** Implement R10 basecalling strategy compatible with garbled sequences prominently found in RBK runs without clean-up.

# In[2]:


# verify running Python version (should be 3.7.5) and adjust jupyter notebook
import IPython
import os
from IPython.core.display import display, HTML      # set display witdth to 100%
display(HTML("<style>.container { width:80% !important; }</style>"))
os.system('python --version')


# In[3]:


# disable extensive logging in jupyter notebook
import logging
logger = logging.getLogger()
logger.setLevel(logging.CRITICAL)


# ## Multithreading and GPU RAM Options
# Depending on the number of parallel threads/cores of the underlying hardware, threading options for multithreaded modules need to be set as environment-specific parameters. One way to do so is through the *os* module.
# 
# GPU RAM is typically limited. Jetson AGX Xavier 32GB should use chunks no larger than 4GB per GPU progress. Guppy basecalling server consumes at least 4GB GPU RAM.

# In[4]:


# execution-wide multithreading options, set according to your hardware. Jetson AGX: suggest "2"
# needs to be set before importing other modules that query these parameters
import os
os.environ["NUMBA_NUM_THREADS"] = "2" # export NUMBA_NUM_THREADS=2
os.environ["OPENBLAS_NUM_THREADS"] = "2" # suggest 2
os.environ["MKL_NUM_THREADS"] = "2"# suggest 2
gpuReservedRam=8*1024**3  # GPU RAM size in bytes, suggest 4GB for Jetson 32GB or 12GB GPU, 8GB for 24GB GPU.


# ## Modules
# This section imports the required modules that should have been installed via pip. Other package managers have not been tested. To install packages, use the setup script provided with this software or, alternatively, install them one by one, ideally in a virtual python environment. Note that the MinKNOW API requires manual patching after installation with pip.

# In[5]:


# Python modules to load
import sklearn # has to be imported before keras first, see https://github.com/keras-team/autokeras/issues/1475
import umap # has to be imported before keras, see https://github.com/keras-team/autokeras/issues/1475
import argparse
import cupy # should be installed even if no GPU is present to avoid code modification when using systems without GPU
import cherrypy
import datetime
import fnmatch
from google.protobuf.json_format import MessageToJson
import logging
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.pyplot import figure
from minknow_api.manager import Manager
import minknow_api.statistics_pb2
import minknow_api.device_pb2
from minknow_api.tools import protocols
from numba import jit
import numpy
import openpyxl
import os
from os import listdir
from pathlib import Path
import pandas
import plotly.express as px
import plotly.graph_objects as go
import psutil
import pysam
import pyslow5
import random
import shutil
import string
import socket
import subprocess
import sys
import time # for development purposes in jupyter notebook (progress bars)
import timeit # benchmarking
from tqdm.notebook import tqdm, trange # for development purposes in jupyter notebook (progress bars)
#import umap # installed via pip install umap-learn; import moved to UMAP function
import webbrowser
from xhtml2pdf import pisa


# ## GPU check
# This code check whether cupy has access to a GPU or not. If not many cupy functions are unavailiable. Not all cupy functions can be managed with the cupy.get_array_module() function. Hence, all instances of cuda calls should be tested individually.

# In[6]:


# check GPU availability, all code below should work in a GPU and non GPU environments; needs to handle cupy and numby differences individually
try:
    cupy.cuda.Device()
    gpu=True
except:
    gpu=False
if gpu:
    import cupy as xp
else:
    import numpy as xp


# ## Configuration
# Below are system-specific parameters that may or may not require adaptation. Many variable names are be self-explanatory. The key difference between Nanopore setups are between devices provided by ONT (MinIT incl. running the MinIT distribution on a NVIDIA Jetson developer kit such as the AGX Xavier, GridION) and the typical Ubuntu-based MinKNOW version on x86_64 computers. The raw data are written into a `/data` directory on ONT-based devices while they are found in `/var/lib/minknow/data` on x86_64 installations. Make sure to adapt your `minknowDataDir` accordingly. There are furthermore permission issues and special folders / files in the MinKNOW data directory. These files / folders should be excluded from analysis through `fileHideList` so that only real run folders will be parsed. Finally, the `nanodipOutputDir` is the place in which the background methylation and alignment process will place its results by replicating the directory hierarchy of the MinKNOW data location. It will not duplicate the data, and these data will be much smaller than raw run data. They can be placed anywhere in the file tree, but also inside the MinKNOW data path within a sub-folder. If the latter is the case, make sure to apply appropriate read/write permissions. Final reports and figures generated by NanoDiP are written into `nanodipReportDir`.

# In[7]:


# configuration parameters, modify here, no need for a configuration file
userhome=str(Path.home())                           # determine home directory (or hardcode it)
gpuid=2                                             # CUDA device ID for cupy operations, e.g., StDev
f5cCudaDevId="1"                                    # CUDA device ID for f5c
nanodipAppDir="/applications/nanodip"               # NanoDiP application directory
nanodipRefDir="/applications/reference_data"        # NanoDiP reference data directory
minknowDataDir="/data"                              # where MinKNOW places its data
epidipTmp="/data/epidip_temp"                       # EpiDiP functionality, temporary file directory
fileHideList=["pings",                              # list of files and folders to be exluded from parsing
              "reads","queued_reads","core-dump-db","lost+found","intermediate","minimap_data","nanodip_tmp","nanodip_output",
              "nanodip_reports","non-ont","raw_for_playback","user_scripts","playback_raw_runs",".Trash-1000",
              "epidip_temp","bismark_output","persistence","recall","dorado_tmp","doradolog","devices","nanodip_basecalling",
              "epidip_CpGs","nanodip_log","api_statistics"]
nanodipOutputDir="/data/nanodip_output" # minknowDataDir+"/nanodip_output"   # location to write intermediate analysis data, i.e. methylation and alignment files
nanodipReportDir="/data/nanodip_reports" # minknowDataDir+"/nanodip_reports"  # location to write reports and figures
nanodipBasecallDir="/data/nanodip_basecalling"      # basecalling intermediate file location for basecalling outside minknow
mk_manager_svc_logpath="/var/log/minknow/mk_manager_svc_log-0.txt" # location of the MinKNOW software log (contains version information, etc.)
seqkits=['SQK-RBK004','SQK-RBK114-24']              # selectable sequencing (rapid bardcoding) kits. First value is considered the default setting.
wantedBases=150000000                               # number of estimated bases until run termination occurs
resultEndings=["_UMAP_top.html","_UMAP_all.html",   # list of file name sections that identify past runs
               "_NanoDiP_report.pdf","_CNVplot.png",
               "_NanoDiP_ranking.pdf"]
analysisExclude=["_foobar_","_TestRun_"] # testing ["_TestRun_"]                       # string patterns in sample names that exclude data from downstream analysis, e.g., test runs
thisFaviconPath=nanodipAppDir+"/favicon.ico"        # the web browser favicon file for this application
epidipLogoPath=nanodipAppDir+"/EpiDiP_Logo_01.png"  # logo bitmap for PDF reports
imgPath=nanodipAppDir                               # the location where image files for the web application are stored
thisHost="localhost"                                # name of the host computer, typically "localhost"
cherrypyHost="localhost"                            # name of the host, typically "localhost" as well
cherrypyPort=8081                                   # port on which the NanoDiP UI will be served
cherrypyThreads=10000                                # number of concurrent threads allowed to CherryPy, decrease in case of performance problems, default = 100
binDir=nanodipRefDir+"/betaEPIC450Kmix_bin"         # location of the preprocessed beta value data
binIndex=binDir+"/index.csv"                        # the index file of the beta value binary files is stored in a CSV files and is generated by the same R script that creates the beta value float binary files
referenceDir=nanodipRefDir+"/reference_annotations" # location of the XLSX files that contain annotations, i.e. reference file collection definitions
methylCutOff=0.35                                   # cut-off for "unmethylated vs methylated" for Illumina array data; also applicable to other methylation data types
topMatch=15                                         # number of reference cases to be shown in subplot including copy number profile links (not advisable >200, plotly will become really slow)
cnvLinkPrefix="http://epidip.usb.ch/sh/apps/allMethylomesCNV08c/" # URL prefix to load PDF with CNV plot for a given Sentrix ID
cnvLinkSuffix="_CNV_IFPBasel_annotations.pdf"       # URL prefix to load PDF with CNV plot 
epidipUmapCoordUrlRoot="http://epidip.usb.ch/epidip_UMAP/" # URL profix to load current EpiDiP coordinates
epidipUmapCoordUrlFiles=["UMAP_25000_CpGs.xlsx", # list of UMAP coordinate files hosted on EpiDiP
                         "UMAP_50000_CpGs.xlsx",
                         "UMAP_75000_CpGs.xlsx"]
chrLengthsFile=nanodipRefDir+"/hg19_cnv/ChrLengths_hg19.tsv"          # contains three columns, A:chrom. strings, B: chrom. lengths, C: offsets
centromereLocationsBed=nanodipRefDir+"/hg19_cnv/hg19.centromere.bed"  # contains the centromere positions for each chromosome
plotlyRenderMode="webgl"                            # default="webgl", alternative "svg" without proper webgl support (e.g., firefox, use "svg"; slower, but does not require GPU)
barcodeNames=["barcode01","barcode02","barcode03",  # barcode strings, currently kit SQK-RBK004, and corresponding R10 kits
              "barcode04","barcode05","barcode06","barcode07","barcode08","barcode09","barcode10","barcode11","barcode12",
              "barcode13","barcode14","barcode15","barcode16","barcode17","barcode18","barcode19","barcode20","barcode21",
              "barcode22","barcode23","barcode24","barcode25","barcode26","barcode27","barcode28","barcode29","barcode30",
              "barcode31","barcode32","barcode33","barcode34","barcode35","barcode36","barcode37","barcode38","barcode39",
              "barcode40","barcode41","barcode42","barcode43","barcode44","barcode45","barcode46","barcode47","barcode48",
              "barcode49","barcode50","barcode51","barcode52","barcode53","barcode54","barcode55","barcode56","barcode57",
              "barcode58","barcode59","barcode60","barcode61","barcode62","barcode63","barcode64","barcode65","barcode66",
              "barcode67","barcode68","barcode69","barcode70","barcode71","barcode72","barcode73","barcode74","barcode75",
              "barcode76","barcode77","barcode78","barcode79","barcode80","barcode81","barcode82","barcode83","barcode84",
              "barcode85","barcode86","barcode87","barcode88","barcode89","barcode90","barcode91","barcode92","barcode93",
              "barcode94","barcode95","barcode96"]
refgenomefa=nanodipRefDir+"/minimap_data/hg19.fa"   # human reference genome
refgenomemmi=nanodipRefDir+"/minimap_data/hg19_nanodip.mmi" # human reference genome minimap2 mmi
ilmncgmapfile=nanodipRefDir+"/microarray/hg19_HumanMethylation450_15017482_v1-2_cgmap.tsv" # Illumina probe names of the 450K array               
f5cBin="export HDF5_PLUGIN_PATH="+userhome+"/.local/hdf5/lib/plugin; /applications/f5c_r10/f5c-v1.2/f5c" # f5cBin="/applications/f5c/f5c" # f5c binary launch command
minimap2Bin="/applications/mm2-ax_versions/mm2-gb/minimap2 -t 1 --gpu-chain --gpu-cfg /applications/mm2-ax_versions/mm2-gb/gpu/orin32GB.json " # "/opt/ont/dorado/bin/minimap2-2.24" #minimap2Bin="/applications/nanopolish/minimap2/minimap2" # minimap2 binary location (absolute path)
samtoolsBin="/applications/samtools/samtools"       # samtools binary location (absolute path)
rscriptBin="/applications/R-4.0.3/bin/Rscript"      # Rscript binary location (absolute path)
readCpGscript=nanodipAppDir+"/readCpGs_mod02.R"     # R script that reads CpGs into simplified text file (absolute path)
cnvGeneListDir=nanodipRefDir+"/cnv_gene_lists/"
cgmapPath="/applications/reference_data/microarray/hg19_HumanMethylation450_15017482_v1-2_cgmap.tsv"
ilmn450kManifestPath=nanodipRefDir+"/microarray/HumanMethylation450_15017482_v1-2.csv"
chromPosPath=nanodipRefDir+"/hg19_cnv/hg19_chromosomes.tsv"
idatPath="/applications/epidip_demo_data/data/demo_idat"
cnvIndexPath="/applications/epidip_demo_data/data/demo_output/CNV_index.csv"
cnvDataPath="/applications/epidip_demo_data/data/demo_output"
cnvHtmlPath="/applications/epidip_demo_data/data/demo_output"
geneWwwPrefix="https://cancer.sanger.ac.uk/cosmic/search?q="
geneWwwSuffix=""
clipVal=1 # color clipping for plot red-green-gradient, +/- value
infiniumBetaRscript="/applications/nanodip/infinium_getBeta_01.R"
inifiniumCnvRscript="/applications/nanodip/infinium_CNV_01.R"
infiniumRscriptBin="/applications/R-4.1.1/bin/Rscript"
infiniumBlacklistCsv=nanodipRefDir+"/microarray/infiniumProbeBlacklist.csv" # probes not to be considered in UMAP plot
terminalBin="/usr/bin/xfce4-terminal"
slow5toolsBin="export HDF5_PLUGIN_PATH="+userhome+"/.local/hdf5/lib/plugin; /applications/slow5tools/slow5tools/slow5tools" # needs to import the location of the VBZ/HDF5 plugins 
basecallWrapperScript="/applications/nanodip/dorado_basecaller_barcorder_call_one_fast5_08.sh"
bcmodelR9="basecallModel=dna_r9.4.1_e8_fast@v3.4&barcodeSet=SQK-RBK004"
bcmodelR10f="basecallModel=dna_r10.4.1_e8.2_400bps_fast@v4.3.0&barcodeSet=SQK-RBK114-24"
bcmodelR10s="basecallModel=dna_r10.4.1_e8.2_260bps_fast@v4.1.0&barcodeSet=SQK-RBK114-24"
minMapQ="20"
doradobin="/applications/ont_dorado_versions/dorado-0.5.0-linux-x64/bin/dorado"
doradomodelpath="/applications/ont_dorado_versions/dorado-0.5.0-models"
bcmodelR9="dna_r9.4.1_e8_fast@v3.4"
bcmodelR10f="dna_r10.4.1_e8.2_400bps_fast@v4.3.0"
bcmodelR10s="dna_r10.4.1_e8.2_260bps_fast@v4.1.0"
barcR9="SQK-RBK004"
barcR10="SQK-RBK114-24"
ontfilesuffixes=['.fast5'] # ['.fast5','.pod5'] # ONT data file suffixes
maxretries=5 # number of attempts to be made on a particular ONT file to be processed
maxcpgqueue=2 #n+1 queues allowed in parallel
maxcnvqueue=0 #n+1 queues allowed in parallel


# ## Degbugging
# The following configuration paramters are intended for debugging purposes. While they can be turned on, they may confuse users as they display detailed python console error messages on the web frontend.

# In[8]:


verbosity=1                                        # 0=low log verbosity, 1=high log verbosity (with timestamps, for benchmarking and debugging)
debugLogging=True                                  # CherryPy debug logging incl. access logs (set True for debugging)


# # No user editable code below
# Do not modify the cells below unless you would like to patch errors or create something new.
# ## Sections
# 1. Generic Functions
# 2. MinKNOW API Functions
# 3. CNV Plotter
# 4. UMAP Methylation Plotter
# 5. EpiDiP Functionality
# 6. User Interface Functions
# 7. Report Generator
# 8. CherryPy Web UI

# ### 1. Generic Functions

# In[9]:


def logpr(v,logstring): # logging funcion that reads verbosity parameter
    if v==1:
        print(str(datetime.datetime.now())+": "+str(logstring))


# In[10]:


def restartNanoDiP():
    cherrypy.engine.restart()


# In[11]:


def forceSymlink(src, dst):
    if os.path.exists(dst):
        if os.path.realpath(src) == dst:
            return
        while os.path.exists(dst): # try to remove and repeat until removal was confirmed
            os.unlink(dst)
    os.symlink(src, dst)


# In[12]:


def getRuns(): # lists run folders from MinKNOW data directory in reverse order based on modif. date
    runFolders=[]
    for r in listdir(minknowDataDir):
        if r not in fileHideList:
            f=minknowDataDir+"/"+r
            if os.path.isdir(f):
                runFolders.append([r,float(os.path.getmtime(f))])
    runFolders.sort(key=lambda row: (row[1], row[0]), reverse=True) # sort based on modif. date
    runFolders=[j.pop(0) for j in runFolders] # remove date column after sorting
    return(runFolders)


# In[13]:


def getPredominantBarcode(sampleName): # adapation to dorado basecalling
    predominantBarcode="undetermined"
    maxbarcode=-1
    try:
        fastaList = [os.path.join(dp, f) for dp, dn, filenames in os.walk(nanodipOutputDir+"/"+sampleName) for f in filenames if os.path.splitext(f)[1] == '.fa']
        barcodeHits=[]
        for b in range(len(barcodeNames)):
            c=0
            #for f in fast5List:
            for f in fastaList:
                if barcodeNames[b] in f:
                    fsize=os.path.getsize(f) # determine fasta file size
                    c+=fsize # sum up file sizes per barcode
            barcodeHits.append(c)
        maxbarcode=max(barcodeHits)
        if maxbarcode>1:
            predominantBarcode=barcodeNames[barcodeHits.index(maxbarcode)]
        else:
            predominantBarcode="undetermined"
    except:
        predominantBarcode="error in barcode determination"
    return predominantBarcode


# In[14]:


def datetimestringnow(): # get current date and time as string to create timestamps
    now = datetime.datetime.now()
    return str(now.year).zfill(4)+str(now.month).zfill(2)+str(now.day).zfill(2)+"_"+str(now.hour).zfill(2)+str(now.minute).zfill(2)+str(now.second).zfill(2)  


# In[15]:


def convert_html_to_pdf(source_html, output_filename): # generate reports
    result_file = open(output_filename, "w+b")         # open output file for writing (truncated binary)
    pisa_status = pisa.CreatePDF(                      # convert HTML to PDF
            source_html,                               # the HTML to convert
            dest=result_file)                          # file handle to recieve result
    result_file.close()                                # close output file
    return pisa_status.err                            # return True on success and False on errors


# In[16]:


def getReferenceAnnotations(): # list all reference annotation files (MS Excel XLSX format)
    referenceAnnotations=[]
    for r in listdir(referenceDir):
        if r.endswith('.xlsx'):
            referenceAnnotations.append(r)    
    return referenceAnnotations


# In[17]:


def writeReferenceDefinition(sampleId,referenceFile): # write the filename of the UMAP reference for the 
    with open(nanodipReportDir+'/'+sampleId+'_selected_reference.txt', 'w') as f: # current run into a text file
        f.write(referenceFile)


# In[18]:


def readReferenceDefinition(sampleId): # read the filename of the UMAP reference for the current sample
    try:
        with open(nanodipReportDir+'/'+sampleId+'_selected_reference.txt', 'r') as f:
            referenceFile=f.read()
    except:
        referenceFile=""
    return referenceFile


# In[19]:


def writeRunTmpFile(sampleId,deviceId):
    with open(nanodipReportDir+'/'+sampleId+'_'+deviceId+'_runinfo.tmp', 'a') as f: # current run into a text file
        try:
            runId=getActiveRun(deviceId)
        except:
            runId="none"
        ro=getThisRunOutput(deviceId,sampleId,runId)
        readCount=ro[0]
        bascalledBases=ro[1]
        overlapCpGs=getOverlapCpGs(sampleId)
        f.write(str(int(time.time()))+"\t"+
                str(readCount)+"\t"+
                str(bascalledBases)+"\t"+
                str(overlapCpGs)+"\n")


# In[20]:


def readRunTmpFile(sampleId):
    print("readRunTmpFile not ready")


# In[21]:


def getOverlapCpGs(sampleName):
    methoverlapPath=nanodipOutputDir+"/"+sampleName # collect matching CpGs from sample
    methoverlapTsvFiles=[] # find all *methoverlap.tsv files
    for root, dirnames, filenames in os.walk(methoverlapPath):
        for filename in fnmatch.filter(filenames, '*methoverlap.tsv'):
            methoverlapTsvFiles.append(os.path.join(root, filename))
    methoverlap=[]
    first=True
    for f in methoverlapTsvFiles:
        try: # some fast5 files do not contain any CpGs
            m=pandas.read_csv(f, delimiter='\t', header=None, index_col=0)
            if first:
                methoverlap=m
                first=False
            else:
                methoverlap=methoverlap.append(m)
        except:
            logpr(verbosity,"empty file encountered, skipping")
    return len(methoverlap)


# In[22]:


def f5cOneFast5(sampleId,refAnno):
    thisRunDir=minknowDataDir+"/"+sampleId
    maxBcCount=0
    targetBc="undetermined"
    allFiles=os.walk(thisRunDir)
    targetBc=getPredominantBarcode(sampleId)
    if refAnno=="predominantbarcode": # only analyze the predominant barcode
        predominantonly=True
    elif refAnno=="includeunclassified":
        predominantonly=False    
    if predominantonly==False:
        targetBcUnclassified='unclassified' # if unclassified reads should be considered as well, append "unclassified" to the query list
        targetBcSet=targetBc+", "+targetBcUnclassified
    else:
        targetBcUnclassified=targetBc #this will essentially search twice for the predominant barcode.
        targetBcSet=targetBc
    fileList = []
    allFiles=os.walk(thisRunDir)
    for dName, sdName, fList in allFiles:
        logpr(verbosity,"dName="+dName)
        if "fast5_pass" in dName:
            if targetBc in dName or targetBcUnclassified in dName:
                l=len(fList)
                if l>1:
                    for f in fList:
                        if f.endswith(".fast5"):
                            logpr(verbosity,dName+"/"+f)
                            fileList.append(dName+"/"+f)   
    nonbarcoded=False  # barcoding might not have been used, in such case the "include unclassified" option can be used to pool all fast5 files
    logpr(verbosity,"targetBc="+targetBc)
    if predominantonly==False:
        logpr(verbosity,"predominantonly==False")
        allFiles=os.walk(thisRunDir) # need to re-walk for re-evaluation
        if targetBc=="undetermined":
            if len(fileList)==0: # no fast5 files have been identified so far, probably from a non-barcoded run
                logpr(verbosity,allFiles)
                for dName, sdName, fList in allFiles:
                    logpr(verbosity,"dName="+dName)
                    if "fast5_pass" in dName:                
                        l=len(fList)
                        if l>1:
                            for f in fList:
                                if f.endswith(".fast5"):
                                    logpr(verbosity,dName+"/"+f)
                                    fileList.append(dName+"/"+f)
                                    nonbarcoded=True
    logpr(verbosity,fileList)
    analyzedCount=0
    calledList=[]
    completedCount=0
    f5cAnalysisDir=nanodipOutputDir+"/"+sampleId
    if os.path.exists(f5cAnalysisDir)==False:
        os.mkdir(f5cAnalysisDir)
    thisBcFast5=[]
    thisBcFastq=[]
    for f in fileList:
        if targetBc in f or targetBcUnclassified in f or nonbarcoded==True:
            q=f.replace(".fast5","").replace("fast5_pass","fastq_pass")+".fastq"
            qgz=f.replace(".fast5","").replace("fast5_pass","fastq_pass")+".fastq.gz" # alternatively look for gzipped fastq
            if os.path.exists(q)==False: # check if accompanying fastq exists
                if os.path.exists(qgz): # check if alternatively, a fastq.gz exists
                    q=qgz # work with qgz path
            if os.path.exists(q): # check if accompanying fastq exists
                thisBcFast5.append(f)
                thisBcFastq.append(q)
                thisBcFileName=f.split("/")
                thisBcFileName=thisBcFileName[len(thisBcFileName)-1].replace(".fast5","") # get name prefix (to be the analysis subdir name later)
                thisAnalysisDir=f5cAnalysisDir+"/"+thisBcFileName
                if os.path.exists(thisAnalysisDir)==False:
                    os.mkdir(thisAnalysisDir)
                target5=thisAnalysisDir+"/"+thisBcFileName+".fast5"
                if q==qgz:
                    targetq=thisAnalysisDir+"/"+thisBcFileName+".fastq.gz"
                else:
                    targetq=thisAnalysisDir+"/"+thisBcFileName+".fastq"
                if os.path.exists(target5)==False:
                    os.symlink(f,target5)             # fast5 symlink
                if os.path.exists(targetq)==False:
                    os.symlink(q,targetq)             #fastq symlink
                if os.path.exists(thisAnalysisDir+"/"+thisBcFileName+"-methoverlapcount.txt")==False:
                    #with open(thisAnalysisDir+"/"+thisBcFileName+"-methoverlapcount.txt", 'w') as dummyfile: # create an empty result file to prevent re-examination
                    #    pass                    
                    if analyzedCount==0:
                        thisRunMinknowDir=minknowDataDir+"/"+sampleId
                        analysisDir=nanodipOutputDir+"/"+sampleId
                        os.makedirs(analysisDir, exist_ok=True)
                        allFiles=os.walk(thisRunMinknowDir)
                        dName=f.replace("/"+thisBcFileName+".fast5","")
                        thisFileName=thisBcFileName #thisFileName=f.replace(".fast5","")
                        thisAnalysisSubdir=analysisDir+"/"+thisFileName
                        os.makedirs(thisAnalysisSubdir,exist_ok=True)
                        forceSymlink(dName+"/"+thisFileName+".fast5",thisAnalysisSubdir+"/"+thisFileName+".fast5")
                        fastqPath="" # reset and search for a matching fastq or fastq.gz file. This file would have been called with MinKNOW.
                        fastqSuffix=".fastq" # default
                        for dN, sdN, fL in allFiles:
                                for fn in fL:
                                    if len(sdN)==0:
                                        #if fn.endswith(".fastq") or fn.endswith(".fastq.gz"):
                                        if thisFileName+".fastq" in fn:
                                            fastqPath=dN+"/"+fn
                                            logpr(verbosity,"Found fastq: "+fastqPath)
                                            if fastqPath.endswith(".fastq"):
                                                fastqSuffix=".fastq"
                                            if fastqPath.endswith(".fastq.gz"):
                                                fastqSuffix=".fastq.gz"                    
                        fastqLinkPath=thisAnalysisSubdir+"/"+thisFileName+fastqSuffix # suffix depends on presence or absende of gz compression
                        logpr(verbosity,"Linking fastq: "+fastqPath+" -> "+fastqLinkPath)
                        forceSymlink(fastqPath,fastqLinkPath)
                        blow5Path=thisAnalysisSubdir+"/"+thisFileName+".blow5"
                        if os.path.exists(blow5Path)==False:
                            cmd=slow5toolsBin+" f2s -p 8 "+thisAnalysisSubdir+" -o "+ blow5Path # convert fast5 to blow5
                            logpr(verbosity,cmd)
                            p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE) 
                            p.wait() # wait until process is completed
                        else:
                            logpr(verbosity,"Path exists: "+blow5Path)
                        analysisFiles=os.walk(thisAnalysisSubdir)
                        for dN, sdN, fL in analysisFiles:
                                for fn in fL:
                                    if len(sdN)==0:
                                        if fn.endswith(".fastq.gz"):
                                            fastqPath=thisAnalysisSubdir+"/"+fn
                        logpr(verbosity,"Found fastq: "+fastqPath)
                        bamSortedPath=thisAnalysisSubdir+"/"+thisFileName+".bam" # sort BAM
                        if os.path.exists(bamSortedPath)==False and os.path.exists(fastqPath+".index")==False and os.path.exists(fastqPath+"index.fai")==False and os.path.exists(fastqPath+"index.gzi")==False:
                            #cmd=minimap2Bin+" -a -x map-ont "+refgenomemmi+" "+fastqPath+" -t 12 | "+samtoolsBin+" sort -T tmp | "+samtoolsBin+" view -bSq "+minMapQ+" >"+bamSortedPath
                            cmd=minimap2Bin+" -ax map-ont "+refgenomemmi+" "+fastqPath+" -t 12 | "+samtoolsBin+" view -b -F0x900 -F 4 | "+samtoolsBin+" sort -T tmp | "+samtoolsBin+" view -bSq "+minMapQ+" >"+bamSortedPath # no duplicate alignments, only aligned reads
                            logpr(verbosity,cmd)
                            p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE) # get sorted BAM (4 threads)
                            p.wait()
                        else:
                            logpr(verbosity,"Path exists: "+bamSortedPath)
                        if os.path.exists(bamSortedPath+".bai")==False: # index BAM
                            cmd=samtoolsBin+" index "+bamSortedPath
                            logpr(verbosity,cmd)
                            p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE) # index BAM
                            p.wait()
                        else:
                            logpr(verbosity,"Path exists: "+bamSortedPath+".bai")
                        if os.path.exists(thisAnalysisSubdir+"/"+thisFileName+".blow5.idx")==False: # f5c index -d [fast5_folder] [read.fastq|fasta] # or f5c index --slow5 [slow5_file] [read.fastq|fasta]
                            cmd=f5cBin+" index --slow5 "+blow5Path+" "+fastqPath # f5c index
                            logpr(verbosity,cmd)
                            p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                            p.wait() # wait until process is completed
                        else:
                            logpr(verbosity,"Path exists: "+thisAnalysisSubdir+"/"+thisFileName+".blow5.idx")
                        methTsvPath=thisAnalysisSubdir+"/"+thisFileName+"-result.tsv"
                        freqTsvPath=thisAnalysisSubdir+"/"+thisFileName+"-freq.tsv"
                        eventTsvPath=thisAnalysisSubdir+"/"+thisFileName+"-events.tsv"
                        if os.path.exists(methTsvPath)==False: # f5c call-methylation -b [reads.sorted.bam] -g [ref.fa] -r [reads.fastq|fasta] > [meth.tsv] #specify --slow5 [slow5_file] to use a slow5 file instead of fast5
                            cmd=f5cBin+" call-methylation --cuda-mem-frac 0.3 -B2000000 -K4000 -b "+bamSortedPath+" -g "+refgenomefa+" -r "+fastqPath+" --slow5 "+blow5Path+" --meth-out-version=1 --cuda-dev-id "+f5cCudaDevId+" > "+methTsvPath
                            logpr(verbosity,cmd)
                            p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                            p.wait() # wait until process is completed
                        else:
                            logpr(verbosity,"Path exists: "+methTsvPath)
                        if os.path.exists(freqTsvPath)==False: # f5c meth-freq -i [meth.tsv] > [freq.tsv]
                            cmd=f5cBin+" meth-freq -c 2.5 -s -i "+methTsvPath+" > "+freqTsvPath
                            logpr(verbosity,cmd)
                            p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                            p.wait() # wait until process is completed
                        else:
                            logpr(verbosity,"Path exists: "+freqTsvPath)
                        cmd=rscriptBin+" "+readCpGscript+" "+thisAnalysisDir+"/"+thisBcFileName+"-freq.tsv "+ilmncgmapfile+" "+thisAnalysisDir+"/"+thisBcFileName+"-methoverlap.tsv "+thisAnalysisDir+"/"+thisBcFileName+"-methoverlapcount.txt"
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                        p.wait()
                        calledList.append(thisBcFileName)
                        analyzedCount+=1
                else:
                    completedCount+=1
    allmethFiles=os.walk(f5cAnalysisDir) # detemine number of overlap CpGs detected so far
    ocpgs=0
    for dName, sdName, fList in allmethFiles:
        for fName in fList:
            if "methoverlapcount" in fName:  
                moc = open(str(dName)+"/"+str(fName), 'r')
                ocpgs+=int(moc.readline())
                moc.close()    
    return "Target = "+targetBcSet+"<br>Methylation called for "+str(calledList)+". "+str(completedCount+analyzedCount)+"/"+str(len(thisBcFast5))+"; found "+str(ocpgs)+" overlap CpGs."


# In[23]:


def basecallOneFast5(sampleId,basecallModel,barcodeSet):
    bclogtxt=""
    cmd=basecallWrapperScript+" "+sampleId+" "+basecallModel+" "+barcodeSet
    p = subprocess.run([basecallWrapperScript,sampleId,basecallModel,barcodeSet], capture_output=True)
    bclogtxt=bclogtxt+str(p.stdout.decode('utf-8'))
    summarySepString="####@####@####"
    logend=""
    if len(bclogtxt)>200:
        logend=bclogtxt[(len(bclogtxt)-200):]
    summarypos=logend.find(summarySepString)
    summarytxt="Basecalling completed or stopped.<hr><br>"
    if summarypos>0:
        summarytxt=logend[summarypos+len(summarySepString):]+"<hr><br>"
    return summarytxt+bclogtxt


# In[24]:


def processOneOntFile(sampleId,refAnno):
    starttime=timeit.default_timer()
    ofrep=""
    errorreport=""
    targetbarc=""
    processedOne=False
    thisRunDir=minknowDataDir+"/"+sampleId # determine ONT file across dataset
    pathList=[]
    fileList=[]
    allFiles=os.walk(thisRunDir)
    for dName, sdName, fList in allFiles:
        for f in fList:
            for suff in ontfilesuffixes:
                if f.endswith(suff):
                    #logpr(verbosity,dName+"/"+f)
                    pathList.append(dName+"/"+f) # avoid splitting later
                    fileList.append(f) # store file names separately
    logpr(verbosity,str(fileList))
    if refAnno=="includeunclassified":
        logpr(verbosity,"**includeunclassified** not yet implemented.")
    else:
        logpr(verbosity,"Including predominant barcode per file.")
        ofrep=ofrep+"<br>Including predominant barcode per file."
    
    for fileIndex in range(len(pathList)):
        logpr(verbosity,pathList[fileIndex])
        logpr(verbosity,fileList[fileIndex])
        thisOntFile=pathList[fileIndex]
        thisFileBasename=fileList[fileIndex]
        for suff in ontfilesuffixes:
            thisFileBasename=thisFileBasename.replace(suff,"")
        thisPath=nanodipOutputDir+"/"+sampleId+"/"+thisFileBasename
        
        if processedOne==False:
            if not os.path.exists(thisPath): # if the directory does not exist
                os.makedirs(thisPath) # make the directory
            thisRetriesCountFile=thisPath+"/"+thisFileBasename+"-retries.txt"
            thisSlow5=thisPath+"/"+thisFileBasename+".blow5"
            thisBam=thisPath+"/"+thisFileBasename+".bam"

            logpr(verbosity,"thisOntFile:"+thisOntFile)
            logpr(verbosity,"thisPath   :"+thisPath)
            logpr(verbosity,"thisSlow5  :"+thisSlow5)
            logpr(verbosity,"thisBam    :"+thisBam)
            ofrep=ofrep+"<br>processing "+thisOntFile

            rcf = open(thisRetriesCountFile, "a") # avoid too numerous attempts to analyze a broken datapoint by logging the number of retries
            rcf.write(str(datetime.datetime.now())+"\n")
            rcf.close()
        rcf= open(thisRetriesCountFile,'r')
        if len(rcf.readlines())>maxretries: # maximum number of retries exceeded
            logpr(verbosity,"Registered >"+str(maxretries)+" attempts to parse "+str(thisOntFile)+", giving up on this file.")
        else:
            try:
                if processedOne==False:
                    if not os.path.exists(thisSlow5):
                        processedOne=True
                        # convert to slow5
                        cmd=slow5toolsBin+" f2s -p 8 "+thisOntFile+" -o "+thisSlow5
                        logpr(verbosity,cmd)
                        try:
                            p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                            p.wait() # wait until process is completed
                            if int(os.stat(thisSlow5).st_size)<10: #something is wrong if the blow 5 is less than 10 bytes long
                                logpr(verbosity,"The file "+thisSlow5+" seems to be corrput, will not try to open it.")
                                raise # throw and error
                            # determine sequencing kit and pore type
                            thisSlow5Content = pyslow5.Open(thisSlow5,'r')
                            barc=str(thisSlow5Content.get_header_value('sequencing_kit')).upper()
                            logpr(verbosity,"Determined kit as "+str(barc)+".")
                            logpr(verbosity,"Does this match "+str(barcR9)+"?")
                            if barc==barcR9:
                                bcmodel=bcmodelR9
                            logpr(verbosity,"Does this match "+str(barcR10)+"?")
                            if barc==barcR10:
                                bcmodel=bcmodelR10f
                        except:
                            logpr(verbosity,"No kit/barcode kit determinable from input file.")
                            barc=barcR10
                            bcmodel=bcmodelR10f
                        logpr(verbosity,"Setting basecalling model to "+bcmodel+" and barcode set to "+barc)
                        ofrep=ofrep+"<br>Setting basecalling model to "+bcmodel+" and barcode set to "+barc
                        # invoke basecaller
                    if not os.path.exists(thisBam):
                        processedOne=True
                        cmd=doradobin+" basecaller --device cuda:"+str(gpuid)+" --batchsize 256 --no-trim --kit-name "+barc+" "+doradomodelpath+"/"+bcmodel+" "+thisOntFile+" > "+thisBam
                        logpr(verbosity,cmd)
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                        p.wait() # wait until process is completed
                        pysam.index(thisBam) #dorado BAM files are non-indexed and contain non-aligned reads
                    samfile = pysam.AlignmentFile(thisBam, "rb", check_sq=False) # since the reads are non-aligned, check_sq=False
                    bcs=[] # list of all barcode entries
                    qn=[] # list of all read IDs
                    qs=[] # list of all sequences
                    for read in samfile.fetch(until_eof=True): # also fetch unaligned reads (all reads are potentially unaligned)
                        qn.append(read.query_name)
                        qs.append(read.query_sequence)
                        try:
                            bcs.append(read.get_tag("BC")) # try to read the BC (barcode) field
                        except:
                            bcs.append("unclassified") # the field does not exist in dorado's BAM files if no barcode was detected
                    samfile.close()
                    bcdict={}
                    for i in bcs:
                        if i in bcdict: bcdict[i]+=1
                        else: bcdict[i]=1
                    print(bcdict)
                    bclist=list(bcdict.keys())
                    print(str(len(bcs))+" "+str(len(qn))+" "+str(len(qs)))
                    for bc in bclist:
                        fastapath=thisPath+"/"+bc+"_reads.fa"
                        if not os.path.exists(fastapath):
                            fastqfile=open(fastapath,"w")
                            for r in range(len(qn)):
                                if bc==bcs[r]:
                                    fastqfile.write(">"+qn[r]+"\n"+qs[r]+"\n")
                            fastqfile.close()

                    bcdict.pop("unclassified")# ignore unclassified reads
                    (targetbarc,bp)=sorted(bcdict.items(), key=lambda x:x[1],reverse=True)[0] # autodetected most prominent barcode
                    logpr(verbosity,targetbarc)

                    thisFasta=thisPath+"/"+targetbarc+"_reads.fa"
                    thisAlnBam=thisPath+"/"+targetbarc+"_reads.bam"
                    tsvSortedPath=thisPath+"/"+targetbarc+"_reads-meth.tsv"
                    tsvFreqSortedPath=thisPath+"/"+targetbarc+"_reads-freq.tsv"
                    methoverlapPath=thisPath+"/"+targetbarc+"_reads-methoverlap.tsv"
                    methoverlapcountPath=thisPath+"/"+targetbarc+"_reads-methoverlapcount.txt"  

                    methylationcalls=tsvFreqSortedPath
                    overlapMethylationFile=methoverlapPath
                    overlapMethylationCountFile=methoverlapcountPath

                    logpr(verbosity,"methylationcalls            :"+str(methylationcalls))
                    logpr(verbosity,"ilmncgmapfile               :"+str(ilmncgmapfile))
                    logpr(verbosity,"overlapMethylationFile      :"+str(overlapMethylationFile))
                    logpr(verbosity,"overlapMethylationCountFile :"+str(overlapMethylationCountFile))
                    if not os.path.exists(thisAlnBam):
                        processedOne=True
                        # align fasta to reference genome
                        cmd=minimap2Bin+" -ax map-ont "+refgenomemmi+" "+thisFasta+" | "+samtoolsBin+" view -b -F0x900 -F 4 | "+samtoolsBin+" sort > "+thisAlnBam
                        logpr(verbosity,cmd)
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                        p.wait() # wait until process is completed
                        pysam.index(thisAlnBam) # index the BAM file        

                        #f5c index
                        cmd=f5cBin+" index --slow5 "+thisSlow5+" "+thisFasta
                        logpr(verbosity,cmd)
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                        p.wait() # wait until process is completed        
                    if not os.path.exists(tsvSortedPath):
                        processedOne=True
                        #5c call methylation"
                        cmd=f5cBin+" call-methylation --cuda-mem-frac 0.3 -B2000000 -K4000 -b "+thisAlnBam+" -g "+refgenomefa+" -r "+thisFasta+" --slow5 "+thisSlow5+" --meth-out-version=1 --cuda-dev-id "+f5cCudaDevId+" > "+tsvSortedPath
                        logpr(verbosity,cmd)
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                        p.wait() # wait until process is completed
                    if not os.path.exists(tsvFreqSortedPath):
                        processedOne=True
                        # f5c meth freq
                        cmd=f5cBin+" meth-freq -c 2.5 -s -i "+tsvSortedPath+" > "+tsvFreqSortedPath
                        logpr(verbosity,cmd)
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                        p.wait() # wait until process is completed
                    if not os.path.exists(methoverlapPath):
                        processedOne=True
                        # read data and reference files
                        ilmncgmap=pandas.read_csv(ilmncgmapfile,header=None,sep="\t")
                        ilmncgmap.columns=["ilmnid","chromosome","strand","start"]
                        display(ilmncgmap)
                        methfreq=pandas.read_csv(methylationcalls,sep="\t")
                        display(methfreq)

                        overlapcpgs=methfreq.merge(ilmncgmap, on=["chromosome","start"])
                        display(overlapcpgs)

                        # determine called singleton CpG sites and write out
                        CpGcalls=overlapcpgs.loc[(overlapcpgs['num_cpgs_in_group']==1) & (overlapcpgs['chromosome']!="chrX") &  (overlapcpgs['chromosome']!="chrY")]
                        display(CpGcalls)

                        CpGout=CpGcalls[['ilmnid', 'called_sites_methylated']].copy()
                        display(CpGout)
                        CpGout.to_csv(methoverlapPath,sep="\t",header=None,index=None)
                    if not os.path.exists(overlapMethylationCountFile):
                        processedOne=True
                        numCpGs=len(CpGout.index)
                        logpr(verbosity,str(numCpGs)+" overlap CpGs found; done!")
                        exectime=str(timeit.default_timer()-starttime)
                        ofrep=ofrep+"<br><b>"+str(numCpGs)+" overlap CpGs found (exec. time: "+exectime+"s)</b><br>"
                        f = open(overlapMethylationCountFile, "w")
                        f.write(str(numCpGs))
                        f.close()
                        # truncate blow5 file to free storage
                        if os.path.isfile(thisSlow5):
                            try:
                                with open(thisSlow5, 'w') as overwriteempty:
                                    pass
                            except:
                                logpr(verbosity,"could not overwrite "+str(thisSlow5))
            except Exception as error:
                ofrep=ofrep+"... error occurred, see report below."
                errorreport="<hr><p>"+str(error)+"</p>"
        if processedOne==False:
            ofrep=ofrep+"... done."
    allmethFiles=os.walk(nanodipOutputDir+"/"+sampleId) # detemine number of overlap CpGs detected so far
    ocpgs=0
    for dName, sdName, fList in allmethFiles:
        for fName in fList:
            if "methoverlapcount" in fName:  
                moc = open(str(dName)+"/"+str(fName), 'r')
                ocpgs+=int(moc.readline())
                moc.close()    
    ofrep="<b>ONT file processing report at "+str(datetime.datetime.now())+"</b><br>Target = "+targetbarc+"; found "+str(ocpgs)+" overlap CpGs.<br>Overall target barcode: "+str(getPredominantBarcode(sampleId))+"<br>"+ofrep+""+errorreport
    return ofrep


# In[25]:


def unique(list1): # function to get unique values (adapted from https://www.geeksforgeeks.org/python-get-unique-values-list/)
    unique_list = []  # initialize a null list 
    for x in list1: # traverse for all elements
        if x not in unique_list: # check if exists in unique_list or not
            unique_list.append(x)
    return unique_list


# In[26]:


def getCnvps(): # lists run folders from MinKNOW data directory in reverse order based on modif. date
    cnvps=[]
    for r in listdir(cnvHtmlPath):
        if r.endswith(".html"):
            f=cnvHtmlPath+"/"+r
            cnvps.append([r,float(os.path.getmtime(f))])
    cnvps.sort(key=lambda row: (row[1], row[0]), reverse=True) # sort based on modif. date
    cnvps=[j.pop(0) for j in cnvps] # remove date column after sorting
    return(cnvps)


# In[27]:


def deleteNonFast5Fastq(deletedir): # delete intermediate data files that are neither fastq nor fast5
    res=""
    for dn,sn,fn in os.walk(deletedir):
        for f in fn:
            if f.endswith(".fastq")==False and f.endswith(".fast5")==False:
                delfile=str(dn)+"/"+str(f)
                os.remove(delfile)
                res=res+delfile+" deleted.<br>"
    return res


# In[28]:


def obtainMethcallStatus(sampleId): # collect number of ONT files, processed ONT files, and overlap CpGs
    allmethFiles=os.walk(nanodipOutputDir+"/"+sampleId) # detemine number of overlap CpGs detected so far
    ocpgs=0
    processedfilecount=0
    for dName, sdName, fList in allmethFiles:
        for fName in fList:
            if "methoverlapcount" in fName:  
                moc = open(str(dName)+"/"+str(fName), 'r')
                ocpgs+=int(moc.readline())
                moc.close()
            if fName.endswith(".blow5"):
                processedfilecount+=1
    ontfilecount=0
    thisRunDir=minknowDataDir+"/"+sampleId # determine ONT file across dataset
    allFiles=os.walk(thisRunDir)
    for dName, sdName, fList in allFiles:
        for f in fList:
            for suff in ontfilesuffixes:
                if f.endswith(suff):
                    ontfilecount+=1
    return [processedfilecount,ontfilecount,ocpgs]   


# ### 2. MinKNOW API Functions
# Check https://github.com/nanoporetech/minknow_api for reference.
# 
# The following code requires a patched version of the MinKNOW API, install it from https://github.com/neuropathbasel/minknow_api.

# In[29]:


def mkManager(): # Construct a manager using the host + port provided. This is used to connect to
    #return Manager(host=thisHost, port=9501, use_tls=False) # the MinKNOW service trough the MK API. changed 20230424
    return Manager(host=thisHost, port=9502) # the MinKNOW service trough the MK API.


# In[30]:


def listMinionPositions(): # list MinION devices that are currenty connected to the system 
    manager = mkManager()
    positions = manager.flow_cell_positions() # Find a list of currently available sequencing positions.  
    return(positions)   # User could call {pos.connect()} here to connect to the running MinKNOW instance.


# In[31]:


def listMinionExperiments(): # list all current and previous runs in the MinKNOW buffer, lost after MinKNOW restart
    manager=mkManager()
    htmlHost="<b>Host: "+thisHost+"</b><br><table border='1'><tr>"
    positions=manager.flow_cell_positions() # Find a list of currently available sequencing positions. 
    htmlPosition=[]
    for p in positions:
        htmlPosinfo="<b>-"+str(p)+"</b><br>"
        connection = p.connect()
        mountedFlowCellID=connection.device.get_flow_cell_info().flow_cell_id # return the flow cell info
        htmlPosinfo=htmlPosinfo+"--mounted flow cell ID: <b>" + mountedFlowCellID +"</b><br>"
        htmlPosinfo=htmlPosinfo+"---"+str(connection.acquisition.current_status())+"<br>" # READY, STARTING, sequencing/mux = PROCESSING, FINISHING; Pause = PROCESSING
        protocols = connection.protocol.list_protocol_runs()
        bufferedRunIds = protocols.run_ids
        for b in bufferedRunIds:
            htmlPosinfo=htmlPosinfo+"--run ID: " + b +"<br>"
            run_info = connection.protocol.get_run_info(run_id=b)
            htmlPosinfo=htmlPosinfo+"---with flow cell ID: " + run_info.flow_cell.flow_cell_id +"<br>"
        htmlPosition.append(htmlPosinfo)
    hierarchy = htmlHost
    for p in htmlPosition:
        hierarchy=hierarchy + "<td valign='top'><tt>"+p+"</tt></td>"
    hierarchy=hierarchy+"</table>"
    return(hierarchy)


# In[32]:


def getFlowCellID(thisDeviceId): # determine flow cell ID (if any). Note that some CTCs have an empty ID string.
    mountedFlowCellID="no_flow_cell"
    manager=mkManager()
    positions=manager.flow_cell_positions() # Find a list of currently available sequencing positions.
    for p in positions:
        if thisDeviceId in str(p):
            connection = p.connect()
            mountedFlowCellID=connection.device.get_flow_cell_info().flow_cell_id # return the flow cell info
    return mountedFlowCellID


# In[33]:


# This cell starts a run on Mk1b devices and perform several checks concerning the run protocol.

# modified from the MinKNOW API on https://github.com/nanoporetech/minknow_api (2021-06)
# adaptation to MinKNOW API 5.4.0 starting Apr 2023, created new fork from ONT's API
# https://github.com/nanoporetech/minknow_api-2
# created from the sample code at
# https://github.com/nanoporetech/minknow_api/blob/master/python/examples/start_protocol.py
# minknow_api.manager supplies "Manager" a wrapper around MinKNOW's Manager gRPC API with utilities
# for querying sequencing positions + offline basecalling tools.
# from minknow_api.manager import Manager

# We need `find_protocol` to search for the required protocol given a kit + product code.
# from minknow_api.tools import protocols
def parse_args():
    """Build and execute a command line argument for starting a protocol.

    Returns:
        Parsed arguments to be used when starting a protocol.
    """
    parser = argparse.ArgumentParser(
        description="""
        Run a sequencing protocol in a running MinKNOW instance.
        """
    )
    parser.add_argument(
        "--host",
        default="localhost",
        help="IP address of the machine running MinKNOW (defaults to localhost)",
    )
    parser.add_argument(
        "--port",
        help="Port to connect to on host (defaults to standard MinKNOW port based on tls setting)",
    )
    parser.add_argument(
        "--no-tls", help="Disable tls connection", default=False, action="store_true"
    )
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging")

    parser.add_argument("--sample-id", help="sample ID to set")
    parser.add_argument(
        "--experiment-group",
        "--group-id",
        help="experiment group (aka protocol group ID) to set",
    )
    parser.add_argument(
        "--position",
        help="position on the machine (or MinION serial number) to run the protocol at",
    )
    parser.add_argument(
        "--flow-cell-id",
        metavar="FLOW-CELL-ID",
        help="ID of the flow-cell on which to run the protocol. (specify this or --position)",
    )
    parser.add_argument(
        "--kit",
        required=True,
        help="Sequencing kit used with the flow-cell, eg: SQK-LSK108",
    )
    parser.add_argument(
        "--product-code",
        help="Override the product-code stored on the flow-cell and previously user-specified"
        "product-codes",
    )
    # BASECALL ARGUMENTS
    parser.add_argument(
        "--basecalling",
        action="store_true",
        help="enable base-calling using the default base-calling model",
    )
    parser.add_argument(
        "--basecall-config",
        help="specify the base-calling config and enable base-calling",
    )
    # BARCODING ARGUMENTS
    parser.add_argument(
        "--barcoding", action="store_true", help="protocol uses barcoding",
    )
    parser.add_argument(
        "--barcode-kits",
        nargs="+",
        help="bar-coding expansion kits used in the experiment",
    )
    parser.add_argument(
        "--trim-barcodes", action="store_true", help="enable bar-code trimming",
    )
    parser.add_argument(
        "--barcodes-both-ends",
        action="store_true",
        help="bar-code filtering (both ends of a strand must have a matching barcode)",
    )

    parser.add_argument(
        "--detect-mid-strand-barcodes",
        action="store_true",
        help="bar-code filtering for bar-codes in the middle of a strand",
    )
    parser.add_argument(
        "--min-score",
        type=float,
        default=0.0,
        help="read selection based on bar-code accuracy",
    )
    parser.add_argument(
        "--min-score-rear",
        type=float,
        default=0.0,
        help="read selection based on bar-code accuracy",
    )

    parser.add_argument(
        "--min-score-mid",
        type=float,
        default=0.0,
        help="read selection based on bar-code accuracy",
    )
    # ALIGNMENT ARGUMENTS
    parser.add_argument(
        "--alignment-reference",
        help="Specify alignment reference to send to basecaller for live alignment.",
    )
    parser.add_argument(
        "--bed-file", help="Specify bed file to send to basecaller.",
    )
    # Output arguments
    parser.add_argument(
        "--fastq",
        action="store_true",
        help="enables FastQ file output, defaulting to 4000 reads per file",
    )
    parser.add_argument(
        "--fastq-reads-per-file",
        type=int,
        default=4000,
        help="set the number of reads combined into one FastQ file.",
    )
    parser.add_argument(
        "--fast5",
        action="store_true",
        help="enables Fast5 file output, defaulting to 4000 reads per file, this will store raw, "
        "fastq and trace-table data",
    )
    parser.add_argument(
        "--fast5-reads-per-file",
        type=int,
        default=4000,
        help="set the number of reads combined into one Fast5 file.",
    )
    parser.add_argument(
        "--bam",
        action="store_true",
        help="enables BAM file output, defaulting to 4000 reads per file",
    )
    parser.add_argument(
        "--bam-reads-per-file",
        type=int,
        default=4000,
        help="set the number of reads combined into one BAM file.",
    )
    # Read until
    parser.add_argument(
        "--read-until-reference", type=str, help="Reference file to use in read until",
    )
    parser.add_argument(
        "--read-until-bed-file", type=str, help="Bed file to use in read until",
    )
    parser.add_argument(
        "--read-until-filter",
        type=str,
        choices=["deplete", "enrich"],
        help="Filter type to use in read until",
    )
    # Experiment
    parser.add_argument(
        "--experiment-duration",
        type=float,
        default=72,
        help="time spent sequencing (in hours)",
    )
    parser.add_argument(
        "--no-active-channel-selection",
        action="store_true",
        help="allow dynamic selection of channels to select pores for sequencing, "
        "ignored for Flongle flow-cells",
    )
    parser.add_argument(
        "--mux-scan-period",
        type=float,
        default=1.5,
        help="number of hours before a mux scan takes place, enables active-channel-selection, "
        "ignored for Flongle flow-cells",
    )
    parser.add_argument(
        "extra_args",
        metavar="ARGS",
        nargs="*",
        help="Additional arguments passed verbatim to the protocol script",
    )
    args = parser.parse_args()
    # Further argument checks
    # Read until must have a reference and a filter type, if enabled:
    if (
        args.read_until_filter is not None
        or args.read_until_reference is not None
        or args.read_until_bed_file is not None
    ):
        if args.read_until_filter is None:
            print("Unable to specify read until arguments without a filter type.")
            sys.exit(1)

        if args.read_until_reference is None:
            print("Unable to specify read until arguments without a reference type.")
            sys.exit(1)

    if args.bed_file and not args.alignment_reference:
        print("Unable to specify `--bed-file` without `--alignment-reference`.")
        sys.exit(1)

    if (args.barcoding or args.barcode_kits) and not (
        args.basecalling or args.basecall_config
    ):
        print(
            "Unable to specify `--barcoding` or `--barcode-kits` without `--basecalling`."
        )
        sys.exit(1)
    if args.alignment_reference and not (args.basecalling or args.basecall_config):
        print("Unable to specify `--alignment-reference` without `--basecalling`.")
        
        sys.exit(1)
    if not (args.fast5 or args.fastq):
        print("No output (fast5 or fastq) specified")

    return args

def is_position_selected(position, args):
    """Find if the {position} is selected by command line arguments {args}."""

    # First check for name match:
    if args.position == position.name:
        return True

    # Then verify if the flow cell matches:
    connected_position = position.connect()
    if args.flow_cell_id is not None:
        flow_cell_info = connected_position.device.get_flow_cell_info()
        if (
            flow_cell_info.user_specified_flow_cell_id == args.flow_cell_id
            or flow_cell_info.flow_cell_id == args.flow_cell_id
        ):
            return True

    return False


def startRun():
    """Entrypoint to start protocol example"""
    # Parse arguments to be passed to started protocols:
    run_id=""
    args = parse_args()
    #args = parse_args(minknowApiShellArgumentString.split())

    # Specify --verbose on the command line to get extra details about
    if args.verbose:
        logging.basicConfig(stream=sys.stdout, level=logging.DEBUG)

    # Construct a manager using the host + port provided:
    #manager = Manager(host=args.host, port=args.port, use_tls=not args.no_tls)
    manager=mkManager()
    errormessage=""
    
    # Find which positions we are going to start protocol on:
    positions = manager.flow_cell_positions()
    filtered_positions = list(
        filter(lambda pos: is_position_selected(pos, args), positions)
    )

    # At least one position needs to be selected:
    if not filtered_positions:
        errormessage="No positions selected for protocol - specify `--position` or `--flow-cell-id`"
    else:
        protocol_identifiers = {}
        for pos in filtered_positions:
            # Connect to the sequencing position:
            position_connection = pos.connect()

            # Check if a flowcell is available for sequencing
            flow_cell_info = position_connection.device.get_flow_cell_info()
            if not flow_cell_info.has_flow_cell:
                errormessage="No flow cell present in position "+str(pos)
            else:
                # Select product code:
                if args.product_code:
                    product_code = args.product_code
                else:
                    product_code = flow_cell_info.user_specified_product_code
                    if not product_code:
                        product_code = flow_cell_info.product_code

                # Find the protocol identifier for the required protocol:
                protocol_info = protocols.find_protocol(
                    position_connection,
                    product_code=product_code,
                    kit=args.kit,
                    basecalling=args.basecalling,
                    basecall_config=args.basecall_config,
                    barcoding=args.barcoding,
                    barcoding_kits=args.barcode_kits,
                )

                if not protocol_info:
                    print("Failed to find protocol for position %s" % (pos.name))
                    print("Requested protocol:")
                    print("  product-code: %s" % args.product_code)
                    print("  kit: %s" % args.kit)
                    print("  basecalling: %s" % args.basecalling)
                    print("  basecall_config: %s" % args.basecall_config)
                    print("  barcode-kits: %s" % args.barcode_kits)
                    print("  barcoding: %s" % args.barcoding)
                    errormessage="Protocol build error, consult application log."
                else:
                    # Store the identifier for later:
                    protocol_identifiers[pos.name] = protocol_info.identifier

                    # Start protocol on the requested postitions:
                    print("Starting protocol on %s positions" % len(filtered_positions))
                    for pos in filtered_positions:

                        # Connect to the sequencing position:
                        position_connection = pos.connect()

                        # Find the protocol identifier for the required protocol:
                        protocol_identifier = protocol_identifiers[pos.name]

                        # Now select which arguments to pass to start protocol:
                        print("Starting protocol %s on position %s" % (protocol_identifier, pos.name))

                        # Set up user specified product code if requested:
                        if args.product_code:
                            position_connection.device.set_user_specified_product_code(
                                code=args.product_code
                            )

                        # Build arguments for starting protocol:
                        basecalling_args = None
                        if args.basecalling or args.basecall_config:
                            barcoding_args = None
                            alignment_args = None
                            if args.barcode_kits or args.barcoding:
                                barcoding_args = protocols.BarcodingArgs(
                                    args.barcode_kits,
                                    args.trim_barcodes,
                                    args.barcodes_both_ends,
                                    args.detect_mid_strand_barcodes,
                                    args.min_score,
                                    args.min_score_rear,
                                    args.min_score_mid,
                                )

                            if args.alignment_reference:
                                alignment_args = protocols.AlignmentArgs(
                                    reference_files=[args.alignment_reference], bed_file=args.bed_file,
                                )

                            basecalling_args = protocols.BasecallingArgs(
                                config=args.basecall_config,
                                barcoding=barcoding_args,
                                alignment=alignment_args,
                            )

                        read_until_args = None
                        if args.read_until_filter:
                            read_until_args = protocols.ReadUntilArgs(
                                filter_type=args.read_until_filter,
                                reference_files=[args.read_until_reference],
                                bed_file=args.read_until_bed_file,
                                first_channel=None,  # These default to all channels.
                                last_channel=None,
                            )

                        def build_output_arguments(args, name):
                            if not getattr(args, name):
                                return None
                            return protocols.OutputArgs(
                                reads_per_file=getattr(args, "%s_reads_per_file" % name)
                            )

                        fastq_arguments = build_output_arguments(args, "fastq")
                        fast5_arguments = build_output_arguments(args, "fast5")
                        bam_arguments = build_output_arguments(args, "bam")

                        # print the protocol parameters
                        print("position_connection "+str(position_connection))
                        print("protocol_identifier "+str(protocol_identifier))
                        print("args.sample_id "+str(args.sample_id))
                        print("args.experiment_group "+str(args.experiment_group))
                        print("basecalling_args "+str(basecalling_args)) 
                        print("read_until_args "+str(read_until_args))
                        print("fastq_arguments "+str(fastq_arguments)) #fastq_arguments OutputArgs(reads_per_file=400)
                        print("fast5_arguments "+str(fast5_arguments)) #fast5_arguments OutputArgs(reads_per_file=400)
                        print("bam_arguments "+str(bam_arguments))
                        print("args.no_active_channel_selection"+str(args.no_active_channel_selection))
                        print("args.mux_scan_period"+str(args.mux_scan_period))
                        print("args.experiment_duration "+str(args.experiment_duration))
                        print("args.extra_args "+str(args.extra_args))  # Any extra args passed.

                        # Now start the protocol:
                        run_id = protocols.start_protocol(
                            position_connection,
                            protocol_identifier,
                            sample_id=args.sample_id,
                            experiment_group=args.experiment_group,
                            basecalling=basecalling_args,
                            read_until=read_until_args,
                            fastq_arguments=fastq_arguments,
                            fast5_arguments=fast5_arguments,
                            bam_arguments=bam_arguments,
                            disable_active_channel_selection=args.no_active_channel_selection,
                            mux_scan_period=args.mux_scan_period,
                            experiment_duration=args.experiment_duration,
                            barcode_info=None, # new in version 5.4.0 20230424: TypeError: start_protocol() missing 1 required positional argument: 'barcode_info'
                            args=args.extra_args,  # Any extra args passed.
                            
                        )

                        #print("Started protocol %s" % run_id)
    return errormessage+run_id # one of them should be ""


# In[34]:


def setBiasVoltage(minionId, newVoltage): # stop an existing run (if any) for a MinION device
    manager=mkManager()
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == minionId, positions))
    # Connect to the grpc port for the position:
    connection = filtered_positions[0].connect()
    v=connection.device.get_bias_voltage().bias_voltage
    thisMessage="Current bias voltage: "+str(v)+" mV. Changing to "+str(newVoltage)+" mV."
    connection.device.set_bias_voltage(bias_voltage=float(newVoltage))
    v=connection.device.get_bias_voltage().bias_voltage
    thisMessage="Voltage after change: "+str(v)+" mV."
    return thisMessage


# In[35]:


def stopRun(minionId): # stop an existing run (if any) for a MinION device
    manager=mkManager()
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == minionId, positions))
    # Connect to the grpc port for the position:
    connection = filtered_positions[0].connect()
    protocols = connection.protocol.list_protocol_runs()
    bufferedRunIds = protocols.run_ids
    thisMessage="No protocol running, nothing was stopped."
    c=0
    for b in bufferedRunIds:
        try:
            connection.protocol.stop_protocol()
            thisMessage="Protocol "+b+" stopped on "+minionId+"."
        except:
            c=c+1
    return thisMessage


# In[36]:


def is_position_selected(position, args): # from minknow_api demos, start_seq.py
    """Find if the {position} is selected by command line arguments {args}."""
    if args.position == position.name: # First check for name match:
        return True
    connected_position = position.connect()  # Then verify if the flow cell matches:
    if args.flow_cell_id is not None:
        flow_cell_info = connected_position.device.get_flow_cell_info()
        if (flow_cell_info.user_specified_flow_cell_id == args.flow_cell_id
            or flow_cell_info.flow_cell_id == args.flow_cell_id):
            return True
    return False


# In[37]:


def getMinKnowApiStatus(deviceString): # MinKNOW detailed status per device
    replyString=""
    manager=mkManager()
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    # determine if anything is running and the kind of run, via set temperature
    # determine if this is a MinION or PromethION instrument; temperature syntax differs
    replyString=replyString+"acquisition.get_acquisition_info().state: "+str(connection.acquisition.get_acquisition_info().state)+"<br>"
    replyString=replyString+"acquisition.current_status(): "+str(connection.acquisition.current_status())+"<br>"
    if getInstrumentType(deviceString)=="MinION":
        replyString=replyString+"minion_device.get_settings().temperature_target.min: "+str(connection.minion_device.get_settings().temperature_target.min)+"<br>"
        replyString=replyString+"device.get_temperature(): " + str(round(connection.device.get_temperature().minion.heatsink_temperature.value,2))+"<br>"
    if getInstrumentType(deviceString)=="P2S":
        replyString=replyString+"promethion_device.get_device_settings().temperature_target: "+str(connection.promethion_device.get_device_settings().temperature_target)+"<br>"
        temps=connection.device.get_temperature().promethion # the P2S has two sensors
        flowcell_temperature=temps.flowcell_temperature.value
        chamber_temperature=temps.chamber_temperature.value
        replyString=replyString+"flowcell_temperature: " + str(round(flowcell_temperature,2))+", chamber_temperature: "+str(round(chamber_temperature,2))+" <br>"
    replyString=replyString+"device.get_bias_voltage(): " + str(connection.device.get_bias_voltage())+"<br>"
    return replyString


# In[38]:


def getMinKnowApiShortStatus(deviceString): # MinKNOW short status information per device
    replyString=""
    manager=mkManager()
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    replyString=replyString+"acquisition.current_status(): "+str(connection.acquisition.current_status())+"<br>"
    return replyString


# In[39]:


def getActiveRun(deviceString): # get handle to an active run on a given MinION device
    manager=mkManager()
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    try:
        activeRun=connection.acquisition.get_current_acquisition_run().run_id # error if no acquisition is running, same as with acquisitio.current_status(), no acquisition until temperature reached
    except:
        activeRun="none"
    return activeRun


# In[40]:


def getRealDeviceActivity(deviceString): # check if a run has been initiated. Earlier versions of NanoDiP (R9 era) used the target temperature
                                         # fails in 2023 with the introduction of R10 and P2S. Replaced by checking
                                         # if there is a sample ID or not
    actrun=getThisRunSampleID(deviceString)
    if "No sampleId information in MinKNOW buffer for" in actrun:
        returnValue="idle"
    else:
        returnValue="sequencing"
    return returnValue


# In[41]:


def getThisRunState(deviceString): # obtain further information about a particular device / run
    manager=mkManager()
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    try:
        thisRunState="Run state for "+deviceString+": "
        thisRunState=thisRunState+str(connection.protocol.get_current_protocol_run().state)+"/"
        thisRunState=thisRunState+str(connection.acquisition.get_acquisition_info().state)
    except:
        thisRunState="No state information in MinKNOW buffer for "+deviceString
    return thisRunState


# In[42]:


def getThisRunSampleID(deviceString): # get SampleID from MinKNOW by device, only available after data
    manager=mkManager()               # acquisition as been initiated by MinKNOW.
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    try:
        thisRunSampleID=connection.protocol.get_current_protocol_run().user_info.sample_id.value
    except:
        thisRunSampleID="No sampleId information in MinKNOW buffer for "+deviceString
    return thisRunSampleID   


# In[43]:


def getThisRunYield(deviceString): # get run yield by device. The data of the previous run will remain 
    manager=mkManager()            # in the buffer until acquisition (not just a start) of a new run
    positions = list(manager.flow_cell_positions()) # have been initiated.
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    try:
        acqinfo=connection.acquisition.get_acquisition_info()
        thisRunYield="Run yield for "+deviceString+"("+acqinfo.run_id+"):&nbsp;"
        thisRunYield=thisRunYield+str(acqinfo.yield_summary)
    except:
        thisRunYield="No yield information in MinKNOW buffer for "+deviceString
    return thisRunYield


# In[44]:


def getThisRunOutput(deviceString,sampleName,runId): # get run yield by device, sampleName, runId
    thisRunOutput=[-1,-1] # defaults in case of error / missing information
    manager=mkManager()            # in the buffer until acquisition (not just a start) of a new run
    positions = list(manager.flow_cell_positions()) # have been initiated.
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    readCount=-3
    calledBases=-3
    if getThisRunSampleID(deviceString)==sampleName: # check that runID and sampleID match
        readCount=-4
        calledBases=-4
        if connection.acquisition.get_current_acquisition_run().run_id==runId:
            if connection.acquisition.current_status()!="status: READY": # i.e., working
                try:
                    acq=connection.acquisition.get_acquisition_info()
                    readCount=acq.yield_summary.basecalled_pass_read_count
                    calledBases=acq.yield_summary.basecalled_pass_bases
                except:
                    readCount=-5
                    calledBases=-5
    thisRunOutput=[readCount,calledBases]
    return thisRunOutput # shall be a list


# In[45]:


def getThisRunEstimatedOutput(deviceString,sampleName,runId): # get run yield by device, sampleName, runId
    thisRunOutput=[-1,-1] # defaults in case of error / missing information
    manager=mkManager()            # in the buffer until acquisition (not just a start) of a new run
    positions = list(manager.flow_cell_positions()) # have been initiated.
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    readCount=-3
    calledBases=-3
    if getThisRunSampleID(deviceString)==sampleName: # check that runID and sampleID match
        readCount=-4
        calledBases=-4
        if connection.acquisition.get_current_acquisition_run().run_id==runId:
            if connection.acquisition.current_status()!="status: READY": # i.e., working
                try:
                    acq=connection.acquisition.get_acquisition_info()
                    readCount=acq.yield_summary.basecalled_pass_read_count
                    calledBases=acq.yield_summary.estimated_selected_bases
                except:
                    readCount=-5
                    calledBases=-5
    thisRunOutput=[readCount,calledBases]
    return thisRunOutput # shall be a list


# In[46]:


def getThisRunInformation(deviceString): # get current run information. Only available after data acquisition
    manager=mkManager()                  # has started.
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position    
    try:
        thisRunInfo="Run information for "+deviceString+"<br><br>"+str(connection.protocol.get_current_protocol_run())
    except:
        thisRunInfo="No protocol information in MinKNOW buffer for "+deviceString
    return thisRunInfo


# In[47]:


def thisRunWatcherTerminator(deviceString,sampleName,wantedBasesOverride): # run terminator site script; periodically reloads and terminates run upon reaching target base count
    realRunId=getActiveRun(deviceString)
    wantedBasesOverride=int(wantedBasesOverride) # convert string to int
    currentBases=getThisRunEstimatedOutput(deviceString,sampleName,realRunId)[1]
    currentBasesString=str(round(currentBases/1e6,2))
    wantedBasesString=str(round(wantedBasesOverride/1e6,2))
    myString="<html><head>"
    myString=myString+"<title>"+currentBasesString+"/"+wantedBasesString+"MB:"+sampleName+"</title>"
    if currentBases < wantedBases: # don't refresh after showing the STOP state
        myString=myString+"<meta http-equiv='refresh' content='10'>"
    myString=myString+"</head><body>"
    myString=myString+"<b>Automatic run terminator</b> for sample <b>"+sampleName+ "</b>, run ID="+realRunId+" on "+deviceString+" when reaching "+wantedBasesString+" MB, now "+currentBasesString+" MB"
    myString=myString+"<hr>"
    myString=myString+"Last refresh at "+datetimestringnow()+".<hr>"
    if currentBases > wantedBasesOverride:
        stopRun(deviceString)
        myString=myString+"STOPPED at "+datetimestringnow()
    elif currentBases==0:
        myString=myString+"IDLE / MUX / ETC"
    else:
        myString=myString+"RUNNING"
    myString=myString+"</body></html>"
    return myString


# In[48]:


def getInstrumentType(deviceString): # determine the instrument type, i.e., promethion, minion, other
    deviceType="other"
    if deviceString.startswith("MN"): # MinION instrument
        deviceType="MinION"
    if deviceString.startswith("P2S"): # PromethION P2 solo (P2S) instrument
        deviceType="P2S"
    return deviceType


# In[49]:


def getRunStatistics(deviceString,runId): # get run statistics like barcode hits for a present or past run
    manager=mkManager()            # in the buffer until acquisition (not just a start) of a new run
    positions = list(manager.flow_cell_positions()) # have been initiated.
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    acqinfo=connection.acquisition.get_acquisition_info()
    readCount=int(acqinfo.yield_summary.basecalled_pass_read_count)
    calledBases=int(acqinfo.yield_summary.basecalled_pass_bases)
    stream = connection.statistics.stream_acquisition_output(
        acquisition_run_id=runId,
        data_selection=minknow_api.statistics_pb2.DataSelection(start=-1,end=-1),
        split=minknow_api.statistics_pb2.AcquisitionOutputSplit(
            alignment_reference=True,
            barcode_name=True,
            lamp_barcode_id=True,
            lamp_target_id=True,
        )
    )
    estimatedSelUnclassBases=0
    estimatedSelBases=pandas.DataFrame(columns=["barcode","estimated_selected_bases"]) # create an empty df
    json_obj_n50=json.loads(MessageToJson(connection.statistics.read_length_n50(acquisition_run_id=runId)))
    estimatedN50=float((json_obj_n50['n50Data'])['estimatedN50'])
    for filter_groups in stream:
        json_obj = json.loads(MessageToJson(filter_groups)) # generates a json string, then parse to json object (dict)
    for s1 in json_obj['snapshots']:
        barcodeName=((s1['filtering'])[0])['barcodeName']
        estimatedSelectedBases=0 # reset
        if len(((s1['snapshots'])[0])['yieldSummary']) > 0: # i.e., there are reads
            estimatedSelectedBases=int((((s1['snapshots'])[0])['yieldSummary'])['estimatedSelectedBases'])
        if barcodeName=="unclassified":
            estimatedSelUnclassBases=estimatedSelectedBases
        else:
            estimatedSelBases.loc[len(estimatedSelBases)] = [barcodeName,estimatedSelectedBases]    
    return [readCount,calledBases,estimatedN50,estimatedSelUnclassBases,estimatedSelBases]


# ### 3. CNV Plotter

# In[50]:


def createCNVPlot(sampleName): # create a genome-wide copy number plot (all-in-one function)
    with tqdm(total=6) as cnvpBar:
        cnvpBar.set_description('CNVP('+sampleName+'), loading reference data')
        cnvpBar.update(1)
        startTime = datetime.datetime.now()
        runPath=nanodipOutputDir+"/"+sampleName
        ChromOffsets = pandas.read_csv(chrLengthsFile, delimiter='\t', header=None, index_col=0)
        validChromosomes=list(ChromOffsets.index)
        ChromOffsetCenters=[]
        for c in range(0,len(validChromosomes)-1):
            ChromOffsetCenters.append((ChromOffsets[2][c]+ChromOffsets[2][c+1])/2)
        lastChromosome=len(validChromosomes)-1
        ChromOffsetCenters.append((ChromOffsets[2][lastChromosome]+ChromOffsets[2][lastChromosome]+ChromOffsets[1][lastChromosome])/2) # last chromosome
        centromereLocations = pandas.read_csv(centromereLocationsBed, delimiter='\t', header=None, index_col=0)
        centromereLocations.loc['chr1'][1]
        centromereOffsets = []
        for c in validChromosomes:
            centromereOffsets.append(ChromOffsets.loc[c][2] + centromereLocations.loc[c][1])
        cnvpBar.set_description('CNVP('+sampleName+'), loading nanopore data')
        cnvpBar.update(1)
        bamFiles=[] # find all bam files
        for root, dirnames, filenames in os.walk(runPath):
            for filename in fnmatch.filter(filenames, '*.bam'):
                bamFiles.append(os.path.join(root, filename))
        cnvScatter=[]
        for thisBam in bamFiles:
            logpr(verbosity,"reading BAM file for CNV: "+str(thisBam))
            try: # try reading the BAM file
                samfile = pysam.AlignmentFile(thisBam, "rb") # pysam coordinates start with 0 while samtools starts with 1 ! See https://pysam.readthedocs.io/en/latest/faq.html#pysam-coordinates-are-wrong
                for thisChromosome in validChromosomes:
                    thisChromOffset=int(ChromOffsets.loc[[thisChromosome]][2])
                    for read in samfile.fetch(thisChromosome):
                        cnvScatter.append(read.pos+thisChromOffset)
            except: # can't read the BAM file
               logpr(verbosity,"can't read BAM file for CNV: "+str(thisBam)) 
        logpr(verbosity,"Number of reads:"+str(len(cnvScatter)))
        #print(max(cnvScatter)/len(cnvScatter))
        cnvpBar.set_description('CNVP('+sampleName+'), determining bin size')
        cnvpBar.update(1)
        binwidth=30*max(cnvScatter)/len(cnvScatter)
        figure(figsize=(20, 3), dpi=120)
        xy=plt.hist(cnvScatter, bins=numpy.arange(min(cnvScatter), max(cnvScatter) + binwidth, binwidth),color="k")
        #plt.vlines(ChromOffsets[2], 0,numpy.max(xy[0]), colors='c', linestyles='solid', label='')
        #plt.title(sampleName, fontdict=None, loc='center', pad=None)
        #plt.yscale('log')
        cnvpBar.set_description('CNVP('+sampleName+'), cleaning data')
        cnvpBar.update(1)
        plotX=xy[1][0:len(xy[0])] # exclude regions with no mapped reads from plot
        plotY=xy[0]
        cleanPlotX=[]
        cleanPlotY=[]
        for p in range(0,len(plotY)):
            if plotY[p]>0:
                cleanPlotX.append(plotX[p])
                cleanPlotY.append(plotY[p])
        cleanPlotX=numpy.array(cleanPlotX) # convert back to numpy array (required for numpy functions)
        cleanPlotY=numpy.array(cleanPlotY)
        yStd=numpy.std(plotY)
        yMean=numpy.mean(plotY)
        yMedian=numpy.median(plotY)
        cleanCoarseX=[]     # local means, cleaned
        cleanCoarseY=[]
        localBinSize=int(10e6)
        localBinStep=int(0.5e6)
        halfLocalBinSize=localBinSize/2
        cnvpBar.set_description('CNVP('+sampleName+'), plotting data')
        cnvpBar.update(1)
        for x in range(0,int(numpy.round(numpy.max(cleanPlotX))),localBinStep):
            thisSlice=cleanPlotY[numpy.logical_and(cleanPlotX >= x,  cleanPlotX <= x+localBinSize)]
            if len(thisSlice)>0:
                cleanCoarseX.append(numpy.median([x,x+localBinSize]))
                cleanCoarseY.append(numpy.median(thisSlice))
        cleanYMedian=numpy.median(cleanPlotY)
        cleanYLower=numpy.min(cleanPlotY)
        cleanYUpper=yMedian*2
        matplotlib.use('Agg')
        figure(figsize=(20, 6), dpi=120)
        plt.ylim(cleanYLower,cleanYUpper)
        plt.scatter(cleanPlotX,cleanPlotY,s=0.2,color='gray',linewidths=1)
        plt.scatter(cleanCoarseX,cleanCoarseY,s=1,linewidths=5,c=cleanCoarseY,cmap=plt.cm.coolwarm_r,vmin=cleanYLower,vmax=cleanYUpper)
        plt.hlines(yMedian, 0, max(cleanPlotX), colors='gray', linestyles='solid', label='') # median line
        plt.vlines(ChromOffsets[2], cleanYLower, cleanYUpper, colors='gray', linestyles='solid', label='')
        plt.vlines(ChromOffsets[2][len(ChromOffsets[2])-1]+ChromOffsets[1][len(ChromOffsets[2])-1], cleanYLower, cleanYUpper, colors='gray', linestyles='solid', label='') # terminating vline
        plt.vlines(centromereOffsets, cleanYLower, cleanYUpper, colors='gray', linestyles='dashed', label='')
        plt.title("Sample ID: "+sampleName, fontdict=None, loc='center', pad=None)
        plt.xlabel('Number of mapped reads: '+str(len(cnvScatter)))
        plt.ylabel('reads per '+ str(round(binwidth/1e6*100)/100) +' MB bin')
        plt.xticks(ChromOffsetCenters, validChromosomes, rotation=90)
        plt.savefig(nanodipReportDir+'/'+sampleName+'_CNVplot.png', bbox_inches='tight')
        readCountFile = open(nanodipReportDir+'/'+sampleName+'_alignedreads.txt',"w")
        readCountFile.write(str(len(cnvScatter)))
        readCountFile.close()
        logpr(verbosity,"CNVP end")
        cnvpBar.set_description('CNVP('+sampleName+'), done')
        cnvpBar.update(1)
        endTime = datetime.datetime.now()
        logpr(verbosity,"Start: "+str(startTime))
        logpr(verbosity,"End  : "+str(endTime))
        logpr(verbosity,"Dur. : "+str(endTime-startTime))


# ### 4. UMAP Methylation Plotter

# In[51]:


def methylationUMAP(sampleName,referenceName): # create a methylation UMAP plot (all-in-one function)
    startTime = datetime.datetime.now()
    logpr(verbosity,"UMAP Plot initiated for "+sampleName)
    with tqdm(total=8) as umapBar:
        umapBar.set_description('UMAP('+sampleName+'), loading annotation')
        umapBar.update(1)
        binFiles=listdir(binDir) # collect reference case binary file names
        referenceString=referenceName.replace(".xlsx","")
        referenceSheetFile=referenceDir+"/"+referenceName # load reference annotation
        referenceSheet=openpyxl.load_workbook(referenceSheetFile)
        referenceList = referenceSheet.active
        col_names = []
        sentrixID  = referenceList['A']
        methClass  = referenceList['B']
        customText = referenceList['C']
        for x in range(3): 
            logpr(verbosity,sentrixID[x].value)
            logpr(verbosity,methClass[x].value)
            logpr(verbosity,customText[x].value)
        indexFile=open(binIndex, "r") # load CpG site index file (contains index for methylation float binary data)
        indexCol=indexFile.read().split("\n")
        indexFile.close()
        umapBar.set_description('UMAP('+sampleName+'), loading and processing methylation data from Nanopore run')
        umapBar.update(1)
        logpr(verbosity,len(indexCol))
        methoverlapPath=nanodipOutputDir+"/"+sampleName # collect matching CpGs from sample
        methoverlapTsvFiles=[] # find all *methoverlap.tsv files
        for root, dirnames, filenames in os.walk(methoverlapPath):
            for filename in fnmatch.filter(filenames, '*methoverlap.tsv'):
                methoverlapTsvFiles.append(os.path.join(root, filename))
        methoverlap=[]
        first=True
        logpr(verbosity,str("Number methoverlap TSV files: "+str(len(methoverlapTsvFiles))))
        for f in methoverlapTsvFiles:
            try: # some fast5 files do not contain any CpGs
                m = pandas.read_csv(f, delimiter='\t', header=None, index_col=0)
                if first:
                    methoverlap = m
                    first = False
                else:
                    methoverlap = methoverlap.append(m)
            except:
                logpr(verbosity,"empty file encountered, skipping")
        logpr(verbosity,str("Number of 450K overlap CpGs: "+str(len(methoverlap))))
        if len(methoverlap)>0:
            overlapProbes=methoverlap.index
            existingProbes=set(indexCol).intersection(overlapProbes) # some probes have been skipped from the reference set, e.g. sex chromosomes
            matching = [indexCol.index(i) for i in existingProbes]
            logpr(verbosity,"overlap probes in cleaned reference data: "+str(len(matching)))
            fileNumbers = []
            binSuffix="_betas_filtered.bin"
            missingFiles=[] # determine if there are entries in the annotation without corresponding methylation binary file
            c=0
            for s in sentrixID:
                try:
                    fileNumbers.append(binFiles.index(s.value+binSuffix))
                except: # e.g. file not available
                    missingFiles.append(c)
                c=c+1
            logpr(verbosity,fileNumbers)
            betaValues=numpy.full([len(matching),len(fileNumbers)],-1, dtype=float, order='C') # create an empty array with -1
            logpr(verbosity,betaValues)
            umapBar.set_description('UMAP('+sampleName+'), loading overlap CpGs from reference data')
            umapBar.update(1)
            matchJumps=numpy.full([len(matching)],-1, dtype=int, order='C')
            matchJumps[0]=matching[0] # create jump list for binary file, add first entry
            if len(matching)>1:
                for m in range(1,len(matching)): # create jump distances for binary file
                    matchJumps[m]=matching[m]-matching[m-1]-1 # concatenate to list
            logpr(verbosity,len(matchJumps))
            betaValues = [ [ None for y in range( len(matching) ) ] for x in range( 1 ) ]
            p_bar = tqdm(range(len(fileNumbers))) # progress bar (development only)
            for f in p_bar:
                betasFilename=binDir+"/"+binFiles[fileNumbers[f]]
                with open(betasFilename, 'rb') as betasFile:
                    allBetaSingleFile = numpy.fromfile(betasFile, dtype=float) # read float with numpy into regular python array (faster) 
                    allBetaSingleFile = numpy.digitize(allBetaSingleFile,bins=[methylCutOff])
                    betaValues.append(allBetaSingleFile[numpy.array(matching)])
                    p_bar.set_description('UMAP('+sampleName+'), loading ref. dataset no. '+str(f))
                betasFile.close()
            umapBar.set_description('UMAP('+sampleName+'), merging reference and nanopore data')
            umapBar.update(1)
            betaValues = numpy.array(betaValues)
            betaValues = numpy.delete(betaValues, 0, 0)
            methoverlapNum=methoverlap.to_numpy()
            diagnosticCaseCgs=[]
            methoverlapCgnames=methoverlap.loc[existingProbes].index # deterine overlap CpG names
            for i in existingProbes:
                thisCg=numpy.mean(methoverlap.loc[[i]].values)
                diagnosticCaseCgs.append(thisCg)
            thisDiagnosticCase=numpy.digitize(diagnosticCaseCgs,bins=[methylCutOff]) # append the nanopore case
            betaValues2=numpy.vstack([betaValues, thisDiagnosticCase]) # convert to numpy array for UMAP function
            del betaValues # free memory
            umapBar.set_description('UMAP('+sampleName+'), calculating embedding')
            umapBar.update(1)
            embeddingAll = umap.UMAP().fit_transform(betaValues2[:,]) # generate UMAP plot
            logpr(verbosity,"\n"+str(embeddingAll))
            umapBar.set_description('UMAP('+sampleName+'), plotting UMAP')
            umapBar.update(1)
            l=len(embeddingAll)-1  # get UMAP coordinates of nanopore case (i.e., last entry in array)
            nanoX=embeddingAll[l,0]
            nanoY=embeddingAll[l,1]
            selectedSentrixIds = [ binFiles[i] for i in fileNumbers]
            logpr(verbosity,len(selectedSentrixIds))
            selectedSentrixIds.append(sampleName)
            logpr(verbosity,len(selectedSentrixIds))
            annoList=[] # create an annotation list and append nanopore case as the last entry
            c=0
            for mc in methClass:
                if c not in missingFiles:
                    annoList.append(mc.value)
                c=c+1
            annoList.append(sampleName)
            embeddingAll=numpy.array(embeddingAll) # convert UMAP data to numpy array
            numberRef=str(len(embeddingAll))+" ref. cases"
            numberCpG=str(len(methoverlap))+" CpGs"
            umapTitle="UMAP for "+sampleName+" against "+referenceName+", "+numberRef+", "+numberCpG
            logpr(verbosity,type(embeddingAll))
            logpr(verbosity,embeddingAll.shape)
            logpr(verbosity,embeddingAll)
            fig2 = px.scatter(x=embeddingAll[:,0], # create UMAP figure with all cases
                              y=embeddingAll[:,1],
                              labels={"x":"UMAP 0",
                                      "y":"UMAP 1"},
                              title=umapTitle, 
                              color=annoList, 
                              hover_name=selectedSentrixIds,
                              render_mode=plotlyRenderMode) #
            fig2.add_annotation(x=nanoX, y=nanoY,
                                text=sampleName,
                                showarrow=True,
                                arrowhead=1)
            fig2.update_yaxes(scaleanchor = "x", scaleratio = 1)
            outPlot=nanodipReportDir+"/"+sampleName+"_"+referenceString+"_UMAP_all.html" # write to HTML file
            fig2.write_html(outPlot)
            fig2.write_image(nanodipReportDir+"/"+sampleName+"_"+referenceString+"_UMAP_all.png")    # plotly image export requires kaleido, install with pip install -U kaleido; needs reloading of plotly to take effect
            umapBar.set_description('UMAP('+sampleName+'), calculating distance ranking')
            umapBar.update(1)
            distances = [] # create distance ranking
            sentrixList = []
            c=0
            for s in sentrixID:
                if c not in missingFiles:
                    sentrixList.append(s.value)
                c=c+1
            sentrixList.append("thisCase")
            mcList = []
            c=0
            for s in methClass:
                if c not in missingFiles:
                    mcList.append(s.value)
                c=c+1
            mcList.append("thisCase")
            txtList = []
            c=0
            for s in methClass:
                if c not in missingFiles:
                    txtList.append(s.value)
                c=c+1
            txtList.append("thisCase")
            caseX=embeddingAll[len(embeddingAll)-1,0]
            caseY=embeddingAll[len(embeddingAll)-1,1]
            xList = []
            yList = []
            for c in embeddingAll:
                distances.append(numpy.sqrt(numpy.power(caseX-c[0],2)+numpy.power(caseY-c[1],2))) # calculate distance
                xList.append(c[0])
                yList.append(c[1])
            distanceRanking = pandas.DataFrame({'distance':distances,'methClass':mcList,'txt':txtList,
                                                'sentrix_ID':sentrixList,'X':xList,'Y':yList})
            distanceRanking = distanceRanking.sort_values(by='distance', axis=0, ascending=True, inplace=False, kind='quicksort')
            # distanceRanking[0:20]
            wb = openpyxl.Workbook()     # write plot coordinates to xlsx
            ws = wb.active # grab the active worksheet
            ws['A1'] = "Sentrix_ID"     # Data can be assigned directly to cells
            ws['B1'] = "X"     
            ws['C1'] = "Y"
            for thisRow in range(len(embeddingAll)):     # Rows can also be appended
                ws.append([selectedSentrixIds[thisRow], embeddingAll[thisRow][0], embeddingAll[thisRow][1]])
            wb.save(nanodipReportDir+"/"+sampleName+"_"+referenceString+"_UMAP.xlsx")     # Save the file
            closeupDf=distanceRanking[0:topMatch] # generate plot of thisCase surrounding reference datasets
            closeupList=closeupDf.values.tolist()
            markerSize=5 # marker size for plotly
            fig3=px.scatter(x=closeupDf["X"],
                            y=closeupDf["Y"],
                            labels={"x":"UMAP 0",
                                    "y":"UMAP 1"},
                            hover_name=closeupDf["sentrix_ID"],
                            title="Close-up "+umapTitle,
                            color=closeupDf["methClass"],
                            render_mode=plotlyRenderMode)
            fig3.update_traces(marker=dict(size=markerSize))
            fig3.add_annotation(x=nanoX, y=nanoY, text=sampleName, showarrow=True, arrowhead=1)
            for ds in closeupDf.values.tolist(): # add transparent hyperlinks to reference CNV plots (e.g., on public EpiDiP.org server)
                fig3.add_annotation(x=ds[4], y=ds[5],
                                    text="<a href='"+cnvLinkPrefix+ds[3]+cnvLinkSuffix+"' target='_blank'>&nbsp;</a>",
                                    showarrow=False, arrowhead=1)
            topRadius=closeupList[len(closeupList)-1][0]
            fig3.add_shape(type="circle",
               x0=nanoX-topRadius,
               y0=nanoY-topRadius,
               x1=nanoX+topRadius,
               y1=nanoY+topRadius,
               line_color="black",
               line_width=0.5)
            fig3.update_yaxes(scaleanchor = "x", scaleratio = 1)
            outPlot=nanodipReportDir+"/"+sampleName+"_"+referenceString+"_UMAP_top.html"
            fig3.write_html(outPlot)
            fig3.write_image(nanodipReportDir+"/"+sampleName+"_"+referenceString+"_UMAP_top.png")
            # create PDF-compatible HTML table with proper cell padding, no borders etc.
            htmlTable="<table border=0>"
            htmlTable+="<tr><td><b>methClass</b></td><td><b>distance</b></td><td><b>txt</b></td><td><b>sentrix_ID</b></td></tr>"
            for i in closeupList:
                htmlTable+="<tr><td>"+str(i[1])+"</td><td>"+str(i[0])+"</td><td>"+str(i[2])+"</td><td>"+str(i[3])+"</td></tr>"
            # generate PDF report
            nanodipReport="<html><head><title>"+sampleName+"</title><body><h1>"+sampleName+"</h1>"+sampleName+"<br>"+htmlTable+"</body>"
            convert_html_to_pdf(nanodipReport, nanodipReportDir+"/"+sampleName+"_"+referenceString+"_NanoDiP_ranking.pdf")
            # generate XLSX version of ranking
            ra = openpyxl.Workbook()     # write plot coordinates to xlsx
            rs = ra.active # grab the active worksheet
            rs['A1'] = "methClass"     # Data can be assigned directly to cells
            rs['B1'] = "distance"     
            rs['C1'] = "txt"
            rs['D1'] = "sentrix_ID"
            for i in closeupList:     # Rows can also be appended
                rs.append([str(i[1]),str(i[0]),str(i[2]),str(i[3])])
            ra.save(nanodipReportDir+"/"+sampleName+"_"+referenceString+"_NanoDiP_ranking.xlsx")     # Save the XLSX file
            cpgCountFile = open(nanodipReportDir+'/'+sampleName+'_cpgcount.txt',"w")
            cpgCountFile.write(str(len(methoverlap)))
            cpgCountFile.close()
            logpr(verbosity,"UMAP end")
            endTime = datetime.datetime.now()
            logpr(verbosity,"Start: "+str(startTime))
            logpr(verbosity,"End  : "+str(endTime))
            logpr(verbosity,"Dur. : "+str(endTime-startTime))
            umapBar.set_description('UMAP('+sampleName+'), completed')
            umapBar.update(1)
        else:
            outPlot=nanodipReportDir+"/"+sampleName+"_"+referenceString+"_UMAP_all.html"
            with open(outPlot, 'w') as txtfile:
                txtfile.write("<html><body>No data to plot.</body></html>")


# ### 5. EpiDiP Functionality
# #### 5.1. EpiDiP UMAP and Reports

# In[52]:


def downloadEpidipCoordinates(umapFile,sentrixid,referenceFile): # download data from EpiDiP server 
    umapLocalFile=nanodipReportDir+"/"+sentrixid+"_"+referenceFile.replace(".xlsx","")+"_UMAP.xlsx"
    cnvLocalFile=nanodipReportDir+"/"+sentrixid+"_CNVplot.pdf"
    p=subprocess.run(["wget", "-O",
                      umapLocalFile,
                      epidipUmapCoordUrlRoot+umapFile],
                      capture_output=True)
    epidipDownloadStatus='exit status: '+str(p.returncode)+'\nstdout: '+str(p.stdout.decode())+'\nstderr: '+str(p.stderr.decode())
    fakeCpgCountFile=nanodipReportDir+"/"+sentrixid+"_cpgcount.txt"
    with open(fakeCpgCountFile, 'w') as f:
        f.write("0")
    f.close()    
    fakeAlignedReadsFile=nanodipReportDir+"/"+sentrixid+"_alignedreads.txt"
    with open(fakeAlignedReadsFile, 'w') as f:
        f.write("0")
    f.close()
    p=subprocess.run(["wget", "-O",
                      cnvLocalFile,
                      cnvLinkPrefix+sentrixid+cnvLinkSuffix],
                      capture_output=True)
    epidipDownloadStatus=epidipDownloadStatus+'<hr>exit status: '+str(p.returncode)+'\nstdout: '+str(p.stdout.decode())+'\nstderr: '+str(p.stderr.decode())
    p=subprocess.run(["pdftoppm","-png","-f","1","-l","1",
                      cnvLocalFile,
                      cnvLocalFile.replace(".pdf","")],
                      capture_output=True)
    epidipDownloadStatus=epidipDownloadStatus+'<hr>exit status: '+str(p.returncode)+'\nstdout: '+str(p.stdout.decode())+'\nstderr: '+str(p.stderr.decode())
    p=subprocess.run(["mv",
                      cnvLocalFile.replace(".pdf","-1.png"),
                      cnvLocalFile.replace(".pdf",".png")],
                      capture_output=True)
    epidipDownloadStatus=epidipDownloadStatus+'<hr>exit status: '+str(p.returncode)+'\nstdout: '+str(p.stdout.decode())+'\nstderr: '+str(p.stderr.decode())
    return str(epidipDownloadStatus)


# In[53]:


def convertXdipCoordinatesToEpidipFormat(umapFile,sentrixid,referenceFile): # convert a locally stored UMAP coordinate file to the format used on the public EpiDiP server
    localUmapFilePath=epidipTmp+"/"+umapFile
    umapLocalFileGenerated=nanodipReportDir+"/"+sentrixid+"_"+referenceFile.replace(".xlsx","")+"_UMAP.xlsx"
    localUmapDf=pandas.read_excel(localUmapFilePath)
    localUmapDf.drop(["Unnamed: 0","MethClass","MethText","binFileName","binPath"],axis=1,inplace=True) # keep: ["SentrixID","UMAP 0","UMAP 1"]
    localUmapDf.columns=["Sentrix_ID","X","Y"]
    localUmapDf.to_excel(umapLocalFileGenerated,index=False)


# In[54]:


def generateEpidipReport(sentrixid, referenceFile, umapFile): # generate PDF report from EpiDiP data (based on case Sentrix ID)    
    downloadLog=downloadEpidipCoordinates(umapFile,sentrixid,referenceFile) # try to download UMAP coordinates, whether they will be used, or not
    if umapFile not in epidipUmapCoordUrlFiles: # this is the case if an xDiP report from a locally computed file was requested. In this case, overwrite the dowloaded UMAP coordinate file with an adapted version of the locally generated coordinate file
        convertXdipCoordinatesToEpidipFormat(umapFile,sentrixid,referenceFile)
    try:
        umapGetScore(sentrixid,referenceFile,sentrixid) # works if the Sentrix ID is valid and data are avaiable
        reportPdfUrl=generatePdfReport(sentrixid,referenceFile,"EpiDiP").replace("'>","' target='_display'>")
        errorlog=""
    except:
        reportPdfUrl=""
        errorlog="""<font color='#ff0000'>The requested data could not be retrieved from the epidip server;
                    search for respective sentrix ID on
                    <a href='http://www.epidip.org' target='_display'>http://www.epidip.org</a>.</font>"""
    res="<b>EpiDiP report: "+reportPdfUrl+"</b>"
    res=res+"<br>Sentrix ID: "+sentrixid
    res=res+"<br>Annotation: "+referenceFile    
    res=res+"<br>UMAP: "+umapFile
    res=res+"<hr><pre>"+downloadLog+"</pre><hr>"
    res=res+errorlog
    return res


# In[55]:


def pool_stats(mempool): # GPU memory allocation checks
    if gpu:
        re='RAM used:'+str(mempool.used_bytes()/1024/1024/1024)+'GB, total:'+str(mempool.total_bytes()/1024/1024/1024)+'GB'
    else:
        re='probably no cuda-compatible device installed'
    return re


# In[56]:


def calculateStdev(referenceName): # calculate sorted standard deviations with GPU (if present) for a particular reference dataset
    if not os.path.exists(epidipTmp):
        os.makedirs(epidipTmp)
    stdFile=epidipTmp+"/"+referenceName+"_stdArray.bin"
    stdSortFile=epidipTmp+"/"+referenceName+"_stdSortArray.bin"
    stdSortFileCsv=epidipTmp+"/"+referenceName+"_stdSortArray.csv"
    binSuffix="_betas_filtered.bin"
    if gpu:
        cupy.cuda.runtime.setDevice(gpuid)
        pool = cupy.cuda.MemoryPool(cupy.cuda.memory.malloc_managed) # get unified pool
        cupy.cuda.set_allocator(pool.malloc) # set unified pool as default allocator
        pool.free_all_blocks() # release GPU memory
    else:
        logpr(verbosity,"probably no CUDA device present.")
    binFiles=listdir(binDir) # collect reference case binary file names
    referenceString=referenceName.replace(".xlsx","")
    referenceSheetFile=referenceDir+"/"+referenceName # load reference annotation
    referenceSheet=openpyxl.load_workbook(referenceSheetFile)
    referenceList = referenceSheet.active
    col_names = []
    sentrixID  = referenceList['A']
    for x in range(3): 
        logpr(verbosity,sentrixID[x].value)
    binSentrix = [w.replace(binSuffix,"") for w in binFiles] # determine which Sentrix IDs from the reference file have corresponding bin files
    sentrixIdPresentList=[]
    binFilePresent=[]
    logpr(verbosity,"length SentrixID "+str(len(sentrixID)))
    p_bar0 = tqdm(range(len(sentrixID)))
    p_bar0.set_description("Determining availablility of binary beta files")
    for x in p_bar0:
        if sentrixID[x].value in binSentrix:
            sentrixIdPresentList.append(sentrixID[x].value)
            binFilePresent.append(binDir+"/"+sentrixID[x].value+binSuffix)
    numFiles=len(sentrixIdPresentList)
    logpr(verbosity,"No. of files for StDev computation: "+str(numFiles))
    betaStdDf=pandas.read_csv(binIndex,header=None,names=["ilmnID"]) # create first column in dataframe, containing ilnmID CpG names
    cpgCount=betaStdDf.shape[0] # rows in ilmnID column
    targetSize=gpuReservedRam #4*1024**3 # desired RAM usage, move to config later # 4 works best on Jetson AGX 32GB; adapt to GPU / RAM layout, see https://forums.developer.nvidia.com/t/nvmapreserveop-0x80000000-failed-22-when-running-cufft-plans/168781/14
    floatSize=8 # size per float in GPU RAM (tested on AGX Xavier)
    chunkSize=round(targetSize/floatSize/numFiles) # determine size of beta value array. Number of CpGs is typically fixed, number of cases is variable   
    if gpu:
        logpr(verbosity,"before array creation: "+str(pool_stats(pool)))
    else:
        logpr(verbosity,"probably no CUDA device present.")
    betaValues=xp.full([numFiles,chunkSize], -1, dtype=float, order='C') # create fixed-size cupy array filled with -1
    if gpu:
        logpr(verbosity,"after array creation: "+str(pool_stats(pool)))
        logpr(verbosity,"cupy array shape: "+str(betaValues.shape))
    else:
        logpr(verbosity,"probably no CUDA device present.")
    cpgChunk=round(cpgCount/chunkSize)
    p_bar1 = tqdm(range(0, cpgCount, chunkSize)) # progress bar
    betaStd=xp.array([])
    c=0
    currentChunkSize=chunkSize
    for x in p_bar1: # break data into chunks, adjusted to GPU RAM availability
        startIndex=x
        endIndex=x+chunkSize-1
        if (endIndex>cpgCount-1): # last round, remaining CpGs
            endIndex=cpgCount-1
            del(betaValues) # delete old array from GPU RAM (dimensions usually won't fit) and create an adjusted one for the last round
            try:
                pool.free_all_blocks() # release GPU memory
            except:
                logpr(verbosity,"probably no CUDA device present.")
            currentChunkSize=endIndex-startIndex+1 
            betaValues=xp.full([numFiles,currentChunkSize], -1, dtype=float, order='C')
        p_bar1.set_description("processing "+str(startIndex)+".."+str(endIndex))
        p_bar2 = tqdm(binFilePresent) # loop through beta value bin files; progress bar
        logpr(verbosity,"currentChunkSize="+str(currentChunkSize)+" betaValues.shape="+str(betaValues.shape))
        c=0 # cupy array index; array is being recycled and overwritten; should not be deleted and re-recreated to save resources and avoid GPU RAM re-allocation
        p_bar2.set_description("CpGs "+str(startIndex)+".."+str(endIndex))
        for f in p_bar2:
            betaValues[c]=xp.fromfile(f,count=currentChunkSize,offset=startIndex, dtype=float)
            c+=1
        betaValues=xp.nan_to_num(betaValues,nan=0.49) # replace nan with 0.49
        logpr(verbosity,"loaded CpG "+str(startIndex)+" to CpG "+str(endIndex))
        betaStd=xp.append(betaStd,xp.std(betaValues, axis=0, dtype=float))
        logpr(verbosity,"betaStd.shape: "+str(betaStd.shape))
    betaStd.tofile(stdFile)
    betaStdDf["binIndex"]=range(0,cpgCount)
    if str(type(betaStd))=="<class 'numpy.ndarray'>":
        betaStdDf["StDev"]=betaStd
    else:
        betaStdDf["StDev"]=betaStd.get() # get is required for cupy arrays only
    try: # probe blacklisting
        ilmnBlacklistedProbes=list(pandas.read_csv(infiniumBlacklistCsv,header=None))
        logpr(verbosity,str(ilmnBlacklistedProbes))
    except:
        ilmnBlacklistedProbes=[]
        logpr(verbosity,"no blacklisted probes")
    display(betaStdDf)
    betaStdDf.replace([numpy.nan,numpy.inf,-numpy.inf],0,inplace=True) # replace NaN and inf with StdDev of 0
    for blp in ilmnBlacklistedProbes: # replace StdDev of blacklisted probes with 0
        betaStdDf.loc[betaStdDf.ilmnID==blp,['StDev']] = 0
    betaStdDf.sort_values(by="StDev", axis=0, ascending=False, inplace=True, kind='quicksort', na_position='last', ignore_index=False, key=None)
    betaStdDf.to_csv(path_or_buf=stdSortFileCsv, index = False)
    del(betaStdDf) # cleanup
    if gpu:
        logpr(verbosity,"cupy array shape: "+str(betaValues.shape))
        logpr(verbosity,"before cleanup: "+str(pool_stats(pool)))
    else:
        logpr(verbosity,"probably no CUDA device present.")
    del(betaStd) # need to release GPU memory explicitly
    del(betaValues) # need to release GPU memory explicitly
    if gpu:
        pool.free_all_blocks() # release GPU memory
        logpr(verbosity,"after cleanup: "+str(pool_stats(pool)))
    else:
        logpr(verbosity,"probably no CUDA device present.")


# In[57]:


def epidipUmap(referenceName, referenceStdName, topN): # calculate UMAP plot from files in a given reference set for topN probes
    import umap
    if not os.path.exists(epidipTmp):
        os.makedirs(epidipTmp)
    stdFile=epidipTmp+"/"+referenceStdName+"_stdArray.bin"
    stdSortFile=epidipTmp+"/"+referenceStdName+"_stdSortArray.bin"
    stdSortFileCsv=epidipTmp+"/"+referenceStdName+"_stdSortArray.csv"
    epidipUmapXlsx=epidipTmp+"/"+datetimestringnow()+"_EpiDiP_"+str(topN)+"_"+referenceName.replace(".xlsx","")+"_"+referenceStdName.replace(".xlsx","")+".xlsx"
    binSuffix="_betas_filtered.bin"
    betaStdDf=pandas.read_csv(stdSortFileCsv,header='infer',sep=",")
    referenceString=referenceName.replace(".xlsx","")
    referenceSheetFile=referenceDir+"/"+referenceName # load reference annotation
    referenceSheet=pandas.read_excel(referenceSheetFile,header=None,names=["SentrixID","MethClass","MethText"])
    binFiles=pandas.DataFrame(listdir(binDir)) # collect reference case binary file names
    binFiles.columns=["binFileName"] # name first column
    binFiles['SentrixID'] = binFiles.apply(lambda row: row.binFileName.replace(binSuffix,""), axis=1) # get SentrixID with string operation on dataframe
    binFiles['binPath'] = binFiles.apply(lambda row: binDir+"/"+row.binFileName, axis=1) # get Path with string operation on dataframe
    referenceSheet=referenceSheet.merge(binFiles, on='SentrixID', how='inner') # get overlap between reference list and available bin files
    numCases=referenceSheet.shape[0]
    floatSize=8 # size per float in GPU RAM (tested on AGX Xavier)
    betaValues=numpy.full([numCases,topN], -1, dtype='float32', order='C') # create fixed-size cupy array filled with -1
    float64bytes=8 # binary float64 representation
    betaStdDf=betaStdDf[0:topN]
    betaStdDf.sort_values(by="binIndex", axis=0, ascending=True, inplace=True, kind='quicksort', na_position='last', ignore_index=False, key=None) # sort the topN lines betaStdDf offsets to facilitate faster loading from permanent storage; rewinding is slow
    betaStdDf['binOffset'] = betaStdDf.apply(lambda row: row.binIndex*float64bytes, axis=1) # pre-calculate offsets (in bytes)
    ind=list(betaStdDf["binOffset"])
    p_bar0 = tqdm(range(numCases))
    c=0
    for f in p_bar0:
        with open(referenceSheet['binPath'][f], "rb") as b:
            buf=bytearray()
            for i in ind:
                b.seek(i) # offset (pre-calculated)
                buf+=b.read(float64bytes) # read bytes into buffer
        betaValues[c]=numpy.float32(numpy.frombuffer(buf,dtype="float64"))
        c+=1
    betaValues=numpy.nan_to_num(betaValues,nan=0.49)
    betaValuesDf=pandas.DataFrame(betaValues)
    embedding=pandas.DataFrame(umap.UMAP().fit_transform(betaValues)) # replace nan with 0.49
    del(betaValues)
    del(betaStdDf)
    embedding.columns=["UMAP 0","UMAP 1"]
    referenceSheet=referenceSheet.join(embedding)
    referenceSheet.to_excel(epidipUmapXlsx)
    return epidipUmapXlsx 


# In[58]:


def plotUmap(epidipUmapXlsx):
    e=pandas.read_excel(epidipUmapXlsx)
    markerSize=5 # marker size for plotly
    u=px.scatter(x=e["UMAP 0"],
                 y=e["UMAP 1"],
                 labels={"x":"UMAP 0",
                         "y":"UMAP 1"},
                 title=epidipUmapXlsx, 
                 color=e['MethClass'], 
                 hover_name=e['SentrixID'],
                 render_mode=plotlyRenderMode)
    u.update_traces(marker=dict(size=markerSize))
    u.update_yaxes(scaleanchor = "x", scaleratio = 1)
    e['cnvlink'] = e.apply(lambda row: "<a href='"+cnvLinkPrefix+row.SentrixID+cnvLinkSuffix+"' target='_blank'>&nbsp;</a>", axis=1)
    u.add_trace(go.Scatter(
        x=list(e['UMAP 0']),
        y=list(e['UMAP 1']),
        mode="text",
        name="CNV links",
        text=list(e['cnvlink']),
        textposition="bottom center",
        hoverinfo="skip",
        visible = "legendonly"))
    u.write_html(epidipUmapXlsx.replace(".xlsx",".html"))


# In[59]:


def plotUmapArrow(epidipUmapXlsx,sentrixList):
    e=pandas.read_excel(epidipUmapXlsx)
    e['cnvlink'] = e.apply(lambda row: "<a href='"+cnvLinkPrefix+row.SentrixID+cnvLinkSuffix+"' target='_blank'>&nbsp;</a>", axis=1)
    markerSize=5 # marker size for plotly
    u=px.scatter(x=e['UMAP 0'],
                 y=e['UMAP 1'],
                 labels={'x':'UMAP 0',
                         'y':'UMAP 1'},
                 title=epidipUmapXlsx, 
                 color=e['MethClass'], 
                 hover_name=e['SentrixID'],
                 render_mode=plotlyRenderMode)
    u.update_traces(marker=dict(size=markerSize))
    u.update_yaxes(scaleanchor = 'x', scaleratio = 1)
    for s in sentrixList:
        i=e[e['SentrixID']==s].index
        if not i.empty:
            u.add_annotation(x=e.loc[i,'UMAP 0'].iat[0], # .iat[0] required to extract respective value, not another pandas dataframe
                             y=e.loc[i,'UMAP 1'].iat[0],
                             text=str(e.loc[i,'SentrixID'].iat[0]),
                             showarrow=True,
                             arrowhead=1)
    u.add_trace(go.Scatter(
        x=list(e['UMAP 0']),
        y=list(e['UMAP 1']),
        mode="text",
        name="CNV links",
        text=list(e['cnvlink']),
        textposition="bottom center",
        hoverinfo="skip",
        visible = "legendonly"))
    uFilename=epidipUmapXlsx.replace(".xlsx",".html")
    u.write_html(uFilename)
    return uFilename.replace(epidipTmp,"")


# In[60]:


def epidipUmapAnnoPage(epidipUmapXlsx,sentrixList):
    if sentrixList==None:
        sentrixList=""
    h='''
        <form action="epidipAnnoUmap" method="GET">
        <table border="0" width="100%" height="90%"><tr>
        <td valign="top" width="25%">
        UMAP coordinates to plot:<br>
        <select name="epidipUmapXlsx">
    '''
    for o in collectUmapXlsx():
        oStr=str(o)
        h=h+'''<option value="'''+oStr+'''">'''+oStr+'''</option>"'''
    h=h+"</select><br><br>"
    h=h+'''
        Sentrix IDs to be marked with arrows:<br>
       <textarea id="sentrixList" name="sentrixList" rows="20" cols="20" placeholder="Enter one more more Sentrix IDs, newline-separated">'''+str(sentrixList)+'''</textarea>
       <br><br>
    '''
    h=h+'''
        <input type="submit" value="show / annotate"/>
        </td>
        <td valign="top" width="75%">
    '''
    if sentrixList!="":
        sentrixList=sentrixList.replace("\r","").split("\n")
        plotHtmlFile=plotUmapArrow(epidipTmp+"/"+epidipUmapXlsx,sentrixList)
        h=h+'''
            <iframe width=99% height=99% src="epidip/'''+str(plotHtmlFile)+'''"></td>
    '''
    else:
        if epidipUmapXlsx!=None:
            h=h+'''
                <iframe width=99% height=99% src="epidip/'''+str(epidipUmapXlsx).replace(".xlsx",".html")+'''"></td>
            '''
        else:
            h=h+'''
                Plot will appear here.<br>'''+str(sentrixList)+'''</td>
            '''
    h=h+'''
            </tr></table>
    '''
    return h


# #### 5.2 EpiDiP CNV Plots

# In[61]:


def absGenomePos(chro,pos,chromPosDict):
    if chro in chromPosDict:
        abspos=pos+chromPosDict[chro]
    else:
        abspos=-1
    return abspos


# In[62]:


def absGenomePosShort(chro,pos,chromPosDictShort):
    if chro in chromPosDictShort:
        abspos=pos+chromPosDictShort[chro]
    else:
        abspos=-1
    return abspos


# In[63]:


def getGeneId(s):
    try:
        l=s.split(";")[0].split("\"")[1]
    except:
        l=""
    return l


# In[64]:


def geneWwwLink(g):
    return "<a href='"+geneWwwPrefix+str(g)+geneWwwSuffix+"' style='color:black;'>"+g+"</a>"


# In[65]:


def cnvClip(v):
    if v>=clipVal:
        v=clipVal
    if v<=-clipVal:
        v=-clipVal
    return v


# In[66]:


def prepareIllArrayReference(): # prepare dataframes containing 450K/EPIC array annotation (for copy number plots) and put them in global variables
    global ilmn450kMinimalDf
    global chromPosDf
    global chromPosDict
    global chromPosDictShort
    cgmapDf=pandas.read_csv(cgmapPath,header=None,sep = '\t')
    cgmapDf.columns=['IlmnID','chromosome','strand','position']
    ilmn450kManifestDf=pandas.read_csv(ilmn450kManifestPath,skiprows=7,low_memory=False)
    ilmn450kMinimalDf=ilmn450kManifestDf[['IlmnID','CHR',"MAPINFO","Strand","UCSC_RefGene_Name"]]
    chromPosDf=pandas.read_csv(chromPosPath,sep="\t")
    absPosList=[]
    for c in range(0,chromPosDf.shape[0]):
        a=0
        for s in range(0,c):
            a+=int(chromPosDf['len'][s])
        absPosList.append(a)
    chromPosDf['absPos'] = absPosList
    chromPosDf.columns=['chrStr', 'len', 'centromere_start', 'centromere_end', 'absPos']
    chromPosDf.rename(columns={"name":"chrString"}) # name "name" causes problems!
    chromPosDf['absCent'] = chromPosDf.apply(lambda row: row.absPos+row.centromere_start,axis=1)
    chromPosDict = dict(zip(list(chromPosDf['chrStr']), list(chromPosDf['absPos'])))
    chromPosDf['chrShort']=chromPosDf.apply(lambda row: row.chrStr.replace('chr',''),axis=1)
    chromPosDictShort = dict(zip(list(chromPosDf['chrShort']), list(chromPosDf['absPos'])))


# In[67]:


def cnvDynamicPlot(mySentrix, searchGenes): # generate a dynamic, annotated, HTML plotly plot
    fitratioIndexPath=cnvDataPath+"/CNV_index.csv"
    fitratioBinPath=cnvDataPath+"/"+mySentrix+"_CNV_fit.bin"
    fitratioSegPath=cnvDataPath+"/"+mySentrix+"_CNV_seg.csv.gz"
    fitratioJoin=pandas.read_csv(fitratioIndexPath,header=None)
    fitratioJoin['ratio']=numpy.fromfile(fitratioBinPath,dtype="float64")
    fitratioJoin.columns=['IlmnID','cnvRatio']
    fitratioJoin=fitratioJoin.merge(ilmn450kMinimalDf, how='right', on = 'IlmnID')
    fitratioJoin['absPos'] = fitratioJoin.apply(lambda row: absGenomePosShort(row.CHR,row.MAPINFO,chromPosDictShort),axis=1)  
    fitratioJoin['cnvValClipped'] = fitratioJoin.apply(lambda row: cnvClip(row.cnvRatio),axis=1)
    fitratioJoin['UCSC_RefGene_List'] = fitratioJoin.apply(lambda row: str(row.UCSC_RefGene_Name) +";",axis=1) # convert NaN to "NaN"
    segDf=pandas.read_csv(fitratioSegPath,compression="gzip") # load conumee segment data
    segDf['absPosStart'] = segDf.apply(lambda row: absGenomePos(row.chrom,row.loc_start,chromPosDict),axis=1)
    segDf['absPosEnd'] = segDf.apply(lambda row: absGenomePos(row.chrom,row.loc_end,chromPosDict),axis=1)
    segDf=segDf.sort_values(by='absPosStart')
    fitratioJoin['UCSC_RefGene_Name'].fillna('',inplace=True) # fill NAs with ''
    p=px.scatter(x=fitratioJoin['absPos'], y=fitratioJoin['cnvRatio'], color=fitratioJoin['cnvValClipped'],
                 hover_name=fitratioJoin['UCSC_RefGene_Name'],
                 color_continuous_scale=["#FF0000", "lightgray","#00FF00"],
                 labels=dict(x="", y="log (intensity)"))
    p.update_layout(title_text="<b>"+mySentrix+"</b>", title_x=0.5)
    p.update_traces(marker=dict(size=4))
    p.update_yaxes(range = [-2,2])
    p.add_hline(y=0, line_width=1, line_dash="dash", line_color="gray")
    for chrom in range(chromPosDf.shape[0]):
        chromStart=chromPosDf['absPos'].iloc[chrom]
        chromCent=chromStart+chromPosDf['centromere_start'].iloc[chrom]
        chromName=chromPosDf['chrStr'].iloc[chrom]
        chromLen=chromPosDf['len'].iloc[chrom]
        p.add_vline(x=chromStart, line_color="gray", line_width=1)
        p.add_vline(x=chromCent, line_color="gray", line_width=1, line_dash="dash")
    p.add_vline(x=chromStart+chromLen, line_color="gray", line_width=1) # closing line for last chromosome
    p.update_xaxes(tickangle=-45,
                   tickmode = 'array',
                   tickvals = chromPosDf['absCent'],
                   ticktext=  chromPosDf['chrStr'])
    for idx,row in segDf.iterrows():
        p.add_shape(type="line",
                    x0=row['absPosStart'], y0=row['seg_mean'],
                    x1=row['absPosEnd'], y1=row['seg_mean'],
                    line=dict(color="blue",width=3))
    for searchGene in searchGenes:
        fitratioJoin['UCSC_RefGene_Name'].fillna('',inplace=True) # fill NAs with ''
        hitGene=fitratioJoin[fitratioJoin['UCSC_RefGene_Name'].str.contains(searchGene+";")] # does not work with NaN values
        hitGene=hitGene.dropna(subset=['cnvRatio', 'absPos','cnvValClipped','UCSC_RefGene_Name'])
        if hitGene.shape[0]>0:
            hitGene['UCSC_RefGene_Name_split'] = hitGene.apply(lambda row: row.UCSC_RefGene_Name.split(";"),axis=1)
            hitGene['Gene_Match'] = hitGene.apply(lambda row: row.UCSC_RefGene_Name_split.count(searchGene),axis=1)
            hitGene=hitGene[hitGene['Gene_Match']>0]        
            if hitGene.shape[0]>0:
                hitGene=hitGene.sort_values(by='absPos')
                p.add_scattergl(x=hitGene['absPos'], y=hitGene['cnvRatio'],marker=dict(color="black", size=5),line=dict(color="black",width=2),name=searchGene)
                hitGeneStart=hitGene['absPos'].min()
                hitGeneEnd=hitGene['absPos'].max()
                hitGeneCnvMed=hitGene['cnvRatio'].median()
                p.add_annotation(x=hitGeneStart,
                                 y=hitGeneCnvMed,
                                 text="<b>"+geneWwwLink(searchGene)+"</b>",
                                 showarrow=True,
                                 arrowhead=5,
                                 font=dict(color="black",size=18))
                p.add_shape(type="line",
                            x0=hitGeneStart, y0=hitGeneCnvMed,
                            x1=hitGeneEnd, y1=hitGeneCnvMed,
                            line=dict(color="magenta",width=3))
    p.update_layout(plot_bgcolor = "#e6ffff")
    p.update_layout(showlegend=False)
    p.update_coloraxes(showscale=False)
    p.write_html(cnvHtmlPath+"/"+mySentrix+"_CNV_dymamicplot.html")    


# In[68]:


def cnvPlotPage(mySentrix,searchGenes):
    if searchGenes:
        searchGeneString=str(searchGenes).replace("[","").replace("]","").replace(" ","").replace("'","")
    else:
        searchGeneString=""
    geneListNames=[]
    geneLists=[]
    for f in listdir(cnvGeneListDir):
        if f.endswith(".txt"):
            if not os.path.isdir(f):
                fi=cnvGeneListDir+"/"+f
                lif=open(fi, "r")
                li=lif.read()
                lif.close()
                geneListNames.append(f)
                geneLists.append(li)
    h='''
        <select id="selecGeneList" onchange="document.getElementById('searchGenes').value = document.getElementById('selecGeneList').value">
        <option value=""> select predefined contents </option>
    '''
    for l in range(len(geneListNames)):
        h=h+'''<option value="'''+geneLists[l]+'''"> '''+geneListNames[l]+''' </option>'''
    h=h+"</select>"    
    h=h+''' 
        <form action="epidipCnv" method="GET">
        <textarea name="searchGenes" id="searchGenes" rows="2" cols="150" placeholder="Enter one more more gene IDs, comma-separated">'''
    h=h+searchGeneString
    h=h+'''</textarea>
        <br>
        <input type="text" name="sentrixId" id="sentrixId" size="20" placeholder="Enter one Sentrix ID">&nbsp;
        <input type="submit" value="create CNV Plot"/>
        </form>
    '''
    if mySentrix and searchGenes:
        try:
            h=h+"CNV Plot for "+str(mySentrix)+", genes: "+str(searchGenes)
            if not "ilmn450kMinimalDf" in globals():
                prepareIllArrayReference()
            searchGenes=str(searchGenes).replace(" ","").split(",")
            cnvDynamicPlot(str(mySentrix),searchGenes)
            h=h+"<br>Plotting completed. Open the CNVP from the table below."
        except Exception as currentError:
            h=h+"<br><font color='red'>Data could not be plotted.<br><pre>"+str(currentError)+"</pre></font>"
    h=h+"<table border=1><tr><td><b>Completed CNV Plots</b></td></tr>"
    for p in getCnvps():
        h=h+"<tr><td><a href='cnv/"+p+"' target='_blank'>"+p+"</a></td></tr>"
    h=h+"</table>"
    return h


# #### 5.3 EpiDiP Microarray Data Preprocessing

# In[69]:


def inifiniumCnvLauncher():
    h="opened "+inifiniumCnvRscript+"<br>"
    cmd=terminalBin+" -T 'IMPORT CNV / close this with ctrl-C' -e '"+infiniumRscriptBin+" "+inifiniumCnvRscript+"'"
    p=subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
    h=h+cmd+"<br><br>You may close this tab now."
    return h


# In[70]:


def inifiniumBetaLauncher():
    h="opened "+infiniumBetaRscript
    cmd=terminalBin+" -T 'IMPORT BETA VALUES / close this with ctrl-C' -e '"+infiniumRscriptBin+" "+infiniumBetaRscript+"'"
    p=subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
    h=h+cmd+"<br><br>You may close this tab now."
    return h


# ### 6. Report Generator (NanoDiP)

# In[71]:


def generatePdfReport(sampleId,referenceId,reportName): # generate a PDF report from either Nanopore or methylation array data
    umapGetScore(sampleId,referenceId, sampleId)
    referenceName=referenceId.replace(".xlsx","")
    runDate=datetimestringnow()
    runSystemName=str(socket.gethostname())
    umapTopPlotPath=nanodipReportDir+"/"+sampleId+"_"+referenceName+"_closeup.png"
    umapScorePlotPath=nanodipReportDir+"/"+sampleId+"_"+referenceName+"_pie.png"
    cnvPlotPath=nanodipReportDir+"/"+sampleId+"_CNVplot.png"
    readCountPath=nanodipReportDir+"/"+sampleId+"_alignedreads.txt"
    cpgCountPath=nanodipReportDir+"/"+sampleId+"_cpgcount.txt"
    reportPdfName=sampleId+"_"+referenceName+"_NanoDiP_report.pdf"
    reportPdfPath=nanodipReportDir+"/"+reportPdfName
    detectedBarcode=getPredominantBarcode(sampleId)
    readCount=open(readCountPath,"r")
    validReads=str(readCount.read())
    readCount.close()
    cpgCount=open(cpgCountPath,"r")
    overlapcpgCount=str(cpgCount.read())
    cpgCount.close()
    htmlcode="""
        <html><body>
        <h1><img src='"""+epidipLogoPath+"""' width='30' align='top'>&nbsp;&nbsp;
        """+reportName+""" Report for Sample """+sampleId+"""</h1>
        <table border='0'>
        <tr><td width='100'>Generated on</td><td width='10'>:</td><td>"""+runSystemName+""" / """+runDate+"""</td></tr>
        <tr><td>Detected barcode</td><td>:</td><td>"""+detectedBarcode+"""</td></tr>
        <tr><td>Valid reads</td><td>:</td><td>"""+validReads+"""</td></tr>
        <tr><td>450K overlap CpG count</td><td>:</td><td>"""+overlapcpgCount+"""</td></tr>
        <tr><td>Reference data</td><td>:</td><td>"""+referenceName+"""</td></tr>
        </table><hr>
        <table border='0'><tr>
        <td valign='top'><b>Methylation UMAP plot</b>
        <img src='"""+umapTopPlotPath+"""' width='350'></td>
        <td valign='top'><b>Methylation UMAP Score</b>
        <img src='"""+umapScorePlotPath+"""' width='350'></td>
        </table><hr>
        <b>Copy number plot</b>
        <img src='"""+cnvPlotPath+"""' width='700'>
        </body><html>
    """
    convert_html_to_pdf(htmlcode, reportPdfPath)
    return "<html><a href='reports/"+reportPdfName+"'>"+"PDF report generated for "+sampleId+", click this link to open it.</a></html>"


# In[72]:


def umapGetScore(sampleId,referenceName,centerId): # get closest neighbor case count from UMAP coordinates and generate respective plots for reporting purposes
    referenceId=referenceName.replace(".xlsx","")
    referenceSheetFile=referenceDir+"/"+referenceName # reference annotation file path
    filePrefix=nanodipReportDir+"/"+sampleId+"_"+referenceId # define input and output files (pathes)
    umapCoordinatesFilePath=filePrefix+"_UMAP.xlsx"
    umapRankingFilePath=filePrefix+"_NanoDiP_ranking.xlsx"
    cpgcountPath=nanodipReportDir+"/"+sampleId+"_cpgcount.txt"
    umapPiePlot=filePrefix+"_pie.png"
    umapPieTxt=filePrefix+"_pie.csv"
    umapDistPlot=filePrefix+"_dist.png"
    umapWholePlot=filePrefix+"_whole.png"
    umapCloseupPlot=filePrefix+"_closeup.png"
    referenceSheet=openpyxl.load_workbook(referenceSheetFile)
    referenceList = referenceSheet.active
    sentrixID  = referenceList['A']
    methClass  = referenceList['B']
    methClassList=[]
    sentrix_IDList=[]
    for r in sentrixID:
        sentrix_IDList.append(r.value)
    for r in methClass:
        methClassList.append(r.value)
    referenceDf=pandas.DataFrame(sentrix_IDList)
    referenceDf.columns=["sentrix_ID"]
    referenceDf['methClass']=methClassList
    referenceCaseCount=len(referenceDf)    
    cpgCountFile = open(cpgcountPath,"r")
    cpgCount = int(cpgCountFile.readline())
    cpgCountFile.close()
    umapCoordinatesFile=openpyxl.load_workbook(umapCoordinatesFilePath) # load pre-calculated UMAP ranking into pandas dataframe
    umapCoordinates=umapCoordinatesFile.active
    umap_sentrix_ID=umapCoordinates['A']
    umap_X=umapCoordinates['B']
    umap_Y=umapCoordinates['C']
    umap_sentrix_ID_list=[]
    umap_X_list=[]
    umap_Y_list=[]
    for r in umap_sentrix_ID:
        umap_sentrix_ID_list.append(str(r.value).replace("_betas_filtered.bin",""))
    for r in umap_X:
        umap_X_list.append(r.value)
    for r in umap_Y:
        umap_Y_list.append(r.value)
    umapDf=pandas.DataFrame(umap_sentrix_ID_list[1:len(umap_sentrix_ID_list)])
    umapDf.columns = ['sentrix_ID']
    umapDf['X']=umap_X_list[1:len(umap_X_list)]
    umapDf['Y']=umap_Y_list[1:len(umap_Y_list)]
    centerCase=umapDf[umapDf['sentrix_ID']==centerId] # select center case by ID aroud which distances shall be determined
    centerCase_X=float(centerCase['X']) # UMAP X of center case
    centerCase_Y=float(centerCase['Y']) # UMAP Y 
    distanceDf=pandas.merge(umapDf,referenceDf,on = 'sentrix_ID')     # merge referenceDf and umapDf to calculate distances
    distanceDf['distance'] = distanceDf.apply(lambda row: numpy.sqrt(numpy.power(centerCase_X-row['X'],2)+numpy.power(centerCase_Y-row['Y'],2)),axis=1) # calculate distances to center case    
    distanceDf = distanceDf.sort_values(by='distance', axis=0, ascending=True, inplace=False, kind='quicksort') # sort distances ascending   
    closeupDf=distanceDf[0:topMatch]
    methClassList=list(closeupDf['methClass'])
    l=len(methClassList)
    hitList=unique(methClassList)
    k = [(x, methClassList.count(x)) for x in set(methClassList)]
    kdf=pandas.DataFrame(k)
    kdf.columns = ['methClass', 'caseCount']
    kdf.sort_values(by=['caseCount'],inplace=True, ascending=False)
    kdf['labelTxt'] = kdf.apply(lambda row: row['methClass']+" ("+str(round(row['caseCount']/l*100))+" %)", axis=1)    
    umapTitle="UMAP for "+sampleId+"<br><sup>reference: "+referenceName+" ("+str(referenceCaseCount)+" cases), "+str(cpgCount)+" CpGs</sup>"
    umapSubtitle=""
    topRadius=float(closeupDf.iloc[[closeupDf.shape[0]-1]]['distance'])
    fig2=px.scatter(x=distanceDf['X'],
                    y=distanceDf['Y'],
                    labels={"x":"UMAP 0",
                            "y":"UMAP 1"},
                    hover_name=distanceDf['sentrix_ID'],
                    title=umapTitle,
                    color=distanceDf['methClass'],
                    width=450,height=400)
    fig2.add_annotation(x=centerCase_X, y=centerCase_Y, text=centerId, showarrow=True, arrowhead=1)
    fig2.add_shape(type="circle",
               x0=centerCase_X-topRadius,
               y0=centerCase_Y-topRadius,
               x1=centerCase_X+topRadius,
               y1=centerCase_Y+topRadius,
               line_color="black",
               line_width=0.5)
    fig2.update_yaxes(scaleanchor = "x", scaleratio = 1)
    fig3=px.scatter(x=closeupDf['X'],
                    y=closeupDf['Y'],
                    labels={"x":"UMAP 0",
                            "y":"UMAP 1",},
                    hover_name=closeupDf['sentrix_ID'],
                    title="Close-up of "+umapTitle,
                    color=closeupDf['methClass'],
                    width=450, height=400)
    fig3.add_annotation(x=centerCase_X, y=centerCase_Y, text=centerId, showarrow=True, arrowhead=1)
    fig3.add_shape(type="circle",
                   x0=centerCase_X-topRadius,
                   y0=centerCase_Y-topRadius,
                   x1=centerCase_X+topRadius,
                   y1=centerCase_Y+topRadius,
                   line_color="black",
                   line_width=0.5)
    fig3.update_yaxes(scaleanchor = "x", scaleratio = 1)
    labelList=[] # get colors from UMAP plot for each label
    colorList=[]
    colorIndex=[]
    for m in range(0,len(fig3['data'])):
        labelList.append(fig3['data'][m]['name'])
        colorList.append(fig3['data'][m]['marker']['color'])
        colorIndex.append(m)
    scatterColorDf=pandas.DataFrame(colorIndex)
    scatterColorDf['methClass']=labelList
    scatterColorDf['colorString']=colorList
    kdf=pandas.merge(kdf,scatterColorDf,on = 'methClass')
    fig1 = px.pie(kdf,values='caseCount',
                  names='labelTxt',
                  color_discrete_sequence=kdf['colorString'],
                  title="Neighbors in "+umapTitle,
                  width=450, height=400)
    fig3.write_image(umapCloseupPlot)
    fig2.write_image(umapWholePlot)
    fig1.write_image(umapPiePlot)
    kdf.to_csv(umapPieTxt, index=True)


# ### 7. User Interface Functions

# In[73]:


def systemStats(): # obtain generic system parameters
    total, used, free = shutil.disk_usage(minknowDataDir)
    m="<tt>"
    m=m+"<b>Software backend versions</b><br>"
    m=m+"<pre>"+readMinknowVersion()+"</pre>"
    m=m+"<br><b>SSD or HDD usage</b><br>"
    m=m+"Total: "+str(total // (2**30))+" GB<br>"
    m=m+"Used : "+str(used // (2**30))+" GB<br>"
    m=m+"Free : "+str(free // (2**30))+" GB<br>"
    m=m+"<br><b>Memory</b><br>"
    m=m+"free : "+str(round(psutil.virtual_memory().available * 100 / psutil.virtual_memory().total))+"%<br>"
    m=m+"<br><b>CPU: </b><br>"    
    m=m+"usage: "+str(round(psutil.cpu_percent()))+"%<br>"
    m=m+"BC&nbsp;&nbsp;&nbsp;active runs: "+str(MinKnowIfPApi.bcQueue)+" <a href='resetQueue?queueName=bc'>reset queue</a><br>"
    m=m+"CpG&nbsp;&nbsp;active runs: "+str(MinKnowIfPApi.cpgQueue)+" <a href='resetQueue?queueName=cpg'>reset queue</a><br>"
    m=m+"CNVP active runs: "+str(MinKnowIfPApi.cnvpQueue)+" <a href='resetQueue?queueName=cnvp'>reset queue</a><br>"
    m=m+"UMAP active runs: "+str(MinKnowIfPApi.umapQueue)+" <a href='resetQueue?queueName=umap'>reset queue</a><br>"
    m=m+"<br><b>GPU</b> active: "+str(gpu)+"<br>"
    m=m+"<br><br>"
    m=m+"<a href='restartServer'>Restart NanoDiP</a> - <font color='#FF0000'><b>USE WITH CAUTION!</b></font>"
    m=m+"</tt>" 
    return m


# In[74]:


def readMinknowVersion(): # read the mk_manager_svc log to revtrieve MinKNOW version information (not through the API but through a log file)
    with open(mk_manager_svc_logpath) as mkmlog:
        mkmlines = mkmlog.readlines()
    versioninfo=""
    st=0
    for li in range(0,30): # check the first 30 lines for system information; varies between MinKNOW versions
        if "hostname" in mkmlines[li]:
            st=li
    for l in range(st+1,st+8): # typically, the following few lines contain useful system information. If output looks non-informative, check the log mkmanager log file and adapt accordingly.
        versioninfo=versioninfo+mkmlines[l][12:]
    return versioninfo


# In[75]:


def analysisLaunchTable(): # present a table from which analyses can be started in a post-hoc manner
    allowedRunFolders=[]
    for a in getRuns():
        inc=True
        for b in analysisExclude:
            if b in a:
                inc=False
        if inc:
            allowedRunFolders.append(a)
    annotations=getReferenceAnnotations()
    lt="<tt><font size='-2'><table border=1>" # lt = launch table, a HTML table with preformed commands to launch analyses
    lt=lt+"<thead><tr><th align='left'><b>Sample ID</b></th>"
    lt=lt+"<th align='left'><b>proc.</b></th><th align='left'><b>total</b></th><th align='left'><b>oCpGs</b></th>"
    lt=lt+"<th align='left'><b>CpGs</b></th><th align='left'><b>CNV</b></th>"
    for a in annotations:
        lt=lt+"<th align='left'><b>UMAP against<br>"+a.replace(".xlsx", "")+"</b></th>"
    lt=lt+"</tr></thead><tbody>"
    for r in range(len(allowedRunFolders)):
        lt=lt+"<tr>"
        lt=lt+"<td>"+allowedRunFolders[r]+"</td>"
        analysisstats=obtainMethcallStatus(allowedRunFolders[r])
        if analysisstats[0]==analysisstats[1]:
            bgcol=" bgcolor='#00ff00'"
        elif analysisstats[0]>analysisstats[1]:
            bgcol=" bgcolor='#ff0000'"
        else:
            bgcol=""
        lt=lt+"<td "+bgcol+">"+str(analysisstats[0])+"</td>"
        lt=lt+"<td>"+str(analysisstats[1])+"</td>"
        lt=lt+"<td>"+str(analysisstats[2])+"</td>"
        lt=lt+"<td><a href='./analysisLauncher?functionName=methylationPoller&sampleName="+allowedRunFolders[r]+"&refAnno=predominantbarcode' target='_blank' rel='noopener noreferrer' title='"+allowedRunFolders[r]+": CpGs with predominant barcode'>get CpGs</a></td>"
        lt=lt+"<td><a href='./analysisLauncher?functionName=cnvplot&sampleName="+allowedRunFolders[r]+"&refAnno=None' target='_blank' rel='noopener noreferrer' title='"+allowedRunFolders[r]+": CNV'>plot CNV</a></td>"
        for a in annotations:
            lt=lt+"<td><a href='./analysisLauncher?functionName=umapplot&sampleName="+allowedRunFolders[r]+"&refAnno="+a+"' target='_blank' rel='noopener noreferrer' title='"+allowedRunFolders[r]+": "+a.replace(".xlsx", "")+"'>plot UMAP</a>&nbsp;"
            lt=lt+"<a href='./makePdf?sampleName="+allowedRunFolders[r]+"&refAnno="+a+"' target='_blank' rel='noopener noreferrer' title='"+allowedRunFolders[r]+": "+a.replace(".xlsx", "")+"'>make PDF</a></td>"    
        lt=lt+"</tr>"
    lt=lt+"</tbody></table></font></tt>"
    return(lt)


# In[76]:


def listRunsTable(): # return all run folders as HTML table
    runFoldersHtml="<table border=1>"
    for r in getRuns():
        runFoldersHtml=runFoldersHtml+"<tr><td>"+r+"</td></tr>"
    runFoldersHtml=runFoldersHtml+"</table>"
    return(runFoldersHtml)


# In[77]:


def collectPastAnalyses(): # list all analysis result files
    fl=[] # empty file file
    for f in listdir(nanodipReportDir):
        for s in resultEndings:
            if s in f:
                fl.append([f,float(os.path.getmtime(nanodipReportDir+"/"+f))])
    fl.sort(key=lambda row: (row[1], row[0]), reverse=True) # sort based on modif. date
    fl=[j.pop(0) for j in fl] # remove date column after sorting
    return fl


# In[78]:


def makePastAnalysesTable(): # create an HTML table displaying links to completed analysis results
    ht="<tt><table border=1>"
    for f in collectPastAnalyses():
        ht=ht+"<tr><td><a href='reports/"+f+"' target='_blank' rel='noopener noreferrer'>"+f+"</a></td></tr>"
    ht=ht+"</tt></table>"
    return ht


# In[79]:


def collectUmapHtml(): # list all UMAP HTML files
    fl=[] # empty file file
    for f in listdir(epidipTmp):
        if "_EpiDiP_" in f:
            if ".html" in f:
                fl.append([f,float(os.path.getmtime(epidipTmp+"/"+f))])
    fl.sort(key=lambda row: (row[1], row[0]), reverse=True) # sort based on modif. date
    fl=[j.pop(0) for j in fl] # remove date column after sorting
    return fl


# In[80]:


def collectUmapXlsx(): # list all UMAP Xlsx files
    fl=[] # empty file file
    for f in listdir(epidipTmp):
        if "_EpiDiP_" in f:
            if ".xlsx" in f:
                fl.append([f,float(os.path.getmtime(epidipTmp+"/"+f))])
    fl.sort(key=lambda row: (row[1], row[0]), reverse=True) # sort based on modif. date
    fl=[j.pop(0) for j in fl] # remove date column after sorting
    return fl


# In[81]:


def makePastUmapTable(): # create an HTML table displaying links completed UMAP results
    ht="<tt><table border=1>"
    for f in collectUmapHtml():
        ht=ht+"<tr><td><a href='epidip/"+f+"' target='_blank' rel='noopener noreferrer'>"+f+"</a></td></tr>"
    ht=ht+"</tt></table>"
    return ht


# In[82]:


def livePage(deviceString): # generate a live preview of the data analysis with the current PNG figures
    thisSampleID=getThisRunSampleID(deviceString) # if there is a run that produces data, the run ID will exist
    thisSampleRef=readReferenceDefinition(thisSampleID).replace(".xlsx", "")
    cnvPlotPath="reports/"+thisSampleID+"_CNVplot.png"
    umapAllPlotPath="reports/"+thisSampleID+"_"+thisSampleRef+"_UMAP_all.png"
    umapAllPlotlyPath="reports/"+thisSampleID+"_"+thisSampleRef+"_UMAP_all.html"
    umapTopPlotPath="reports/"+thisSampleID+"_"+thisSampleRef+"_UMAP_top.png"
    ht="<html><body><tt>sample ID: "+thisSampleID+" with reference "+thisSampleRef+"</tt><br>"
    ht=ht+"<a href='"+cnvPlotPath+"' target='_blank'><img align='Top' src='"+cnvPlotPath+"' width='50%' alt='CNV plot will appear here'></a>"
    ht=ht+"<a href='"+umapAllPlotlyPath+"' target='_blank'><img align='Top' src='"+umapAllPlotPath+"' width='50%' alt='UMAP plot will appear here'></a>"
    ht=ht+"</tt></table><body></html>"
    return ht


# In[83]:


def methcallLivePage(sampleName,refAnno): # generate a self-refreshing page to invoke methylation calling
    ht="<html><head><title>MethCaller: "+sampleName+"</title>"
    ht=ht+"<meta http-equiv='refresh' content='3'></head><body>"
    ht=ht+"last refresh and console output at "+datetimestringnow()+"<hr>shell output<br><br><tt>"
    #ht=ht+calculateMethylationAndBamFromFast5Fastq(sampleName)
    #ht=ht+f5cOneFast5(sampleName,refAnno)
    ht=ht+processOneOntFile(sampleName,refAnno)
    ht=ht+"</tt></body></html>"
    return ht


# In[84]:


def basecallLivePage(sampleName,basecallModel,barcodeSet): # generate a self-refreshing page to invoke basecalling
    ht="<html><head><title>basecaller: "+sampleName+"</title>"
    ht=ht+"<meta http-equiv='refresh' content='3'></head><body>"
    ht=ht+"last refresh and console output at "+datetimestringnow()+"<hr>shell output<br><br><tt>"
    ht=ht+basecallOneFast5(sampleName,basecallModel,barcodeSet)
    ht=ht+"</tt></body></html>"
    return ht


# In[85]:


def menuheader(n,r): # generate a universal website header for the UI pages that contains a simple main menu
    menuList=[["index","Overview","General system information"],
              ["listPositions","Mk1b Status","Live status of all connected Mk1b devices"],
              ["startSequencing","Start seq.","Start a sequencing run on an idle Mk1b device"],
              ["startTestrun","Start test run","Start a test seq. run on an idle Mk1b device to verify that the previous flow cell wash was successful."],
              ["listExperiments","Seq. runs","List all buffered runs. Will be purged upon MinKNOW backend restart."],
              ["listRuns","Results","List all completed analysis results"],
              ["analyze","Analyze","Launch data analyses manually, e.g. for retrospective analysis"],
              ["epidipGenUmap","EpiDiP UMAP","re-generate EpiDiP UMAP plot"],
              ["epidipAnnoUmap","EpiDiP Annotate","annotate EpiDiP UMAP plot"],
              ["epidipCnv","EpiDiP CNVP","create dynamic annotated copy number variation plots"],
              ["epidipReport","EpiDiP report","Generate report PDFs for EpiDiP cases"],
              ["xdipReport","xDiP report","Generate report PDFs for locally generated UMAP coordinates"],
              ["cleanup","Clean-up","Selectively delete intermediate result data"],
              ["about","About NanoDiP","Version, etc."]
             ]
    mm="<html><head><title>NanoDiP Version "+versionString+"</title>"
    if r>0: # autorefresh wanted
        mm=mm+"<meta http-equiv='refresh' content='"+str(r)+"'>"
    mm=mm+'''</head><body><table border=0 cellpadding=2><tr>
    <td><img src='img/EpiDiP_Logo_01.png' width='40px' height='40px'></td>'''
    nn=0
    for m in menuList:
        if n==nn:
            selectedColor="#E0E0E0"
        else:
            selectedColor="white"
        mm=mm+"<td bgcolor='"+selectedColor+"'><b><a href='"+m[0]+"' title='"+m[2]+"'>"+m[1]+"</a></b></td>"
        nn=nn+1
    mm=mm+"</tr></table><br>"
    return(mm)


# In[86]:


def cleanuptable(deltype,deldir): # present table to delete intermediate data
    ct=""
    if deltype and deldir: # there is something destined to deletion
        if deltype=="nanodip_basecalling":
            delpath=nanodipBasecallDir+"/"+deldir
        elif deltype=="nanodip_output":
            delpath=nanodipOutputDir+"/"+deldir
            delsymlinkpath=minknowDataDir+"/"+deldir+"_ND"
            shutil.rmtree(delsymlinkpath, ignore_errors=True, onerror=None)
            ct=ct+delsymlinkpath+" deleted.<br><br>"
            delndpath=minknowDataDir+"/"+deldir
            ct=ct+deleteNonFast5Fastq(delndpath)
        shutil.rmtree(delpath, ignore_errors=True, onerror=None)
        ct=ct+delpath+" deleted.<br><br>"
        
        ct=ct+"<br>"
    ct=ct+"<h1>Delete basecalling data</h1>"
    ct=ct+"<table border=1>"
    ct=ct+"<tr>"
    ct=ct+"<td><tt><b>directory</b></tt></td>"
    ct=ct+"<td><tt><b>delete basecalling data</b></tt></td>"
    ct=ct+"</tr>"
    basecallDirs=[]
    for basecallDir in os.listdir(nanodipBasecallDir):
        if os.path.isdir(os.path.join(nanodipBasecallDir, basecallDir)):
            basecallDirs.append(basecallDir)
    basecallDirs.sort()
    for basecallDir in basecallDirs:
            ct=ct+"<tr><td><tt>"+str(basecallDir)+"</tt></td>"
            ct=ct+"<td><tt><a href='./cleanup?deltype=nanodip_basecalling&deldir="+str(basecallDir)+"'>delete basecalling data</a></tt></td>"
            ct=ct+"</tr>"
    ct=ct+"</table>"
    ct=ct+"<h1>Delete methylation calling data</h1>"
    ct=ct+"<table border=1>"
    ct=ct+"<tr>"
    ct=ct+"<td><tt><b>directory</b></tt></td>"
    ct=ct+"<td><tt><b>delete methylation calling data</b></tt></td>"
    ct=ct+"</tr>"
    outputDirs=[]
    for outputDir in os.listdir(nanodipOutputDir):
        if os.path.isdir(os.path.join(nanodipOutputDir, outputDir)):
            outputDirs.append(outputDir)
    outputDirs.sort()
    for outputDir in outputDirs:
            ct=ct+"<tr><td><tt>"+str(outputDir)+"</tt></td>"           
            ct=ct+"<td><tt><a href='./cleanup?deltype=nanodip_output&deldir="+str(outputDir)+"'>delete methylation calling data</a></tt></td>"
            ct=ct+"</tr>"
    ct=ct+"</table>"
    return ct


# ### 8. CherryPy Web UI
# The browser-based user interface is based on CherryPy, which contains an intergrated web server and serves pages locally. Communication between the service and browser typically generates static web pages that may or may not contain automatic self refresh commands. In the case of self-refreshing pages, the browser will re-request a given page with leads to re-execution of the respective python functions. The main handles to these function are located in the Web UI cell below.

# In[87]:


# Launch CherryPy Webserver. Relaunch this cell only unless other cells have been modified to restart.
class MinKnowIfPApi(object): # the CherryPy Web UI class that defines entrypoints and function calls
    # global variables within the CherryPy Web UI
    globalRunFolders=getRuns()
    globalRunStatus= [None] * len(globalRunFolders)
    cpgQueue=0
    umapQueue=0
    cnvpQueue=0
    bcQueue=0
    
    @cherrypy.expose
    def index(self):
        myString=menuheader(0,15)
        myString=myString+"<tt><b>Computer:</b> "+str(socket.gethostname())+"</tt><br><br>"
        myString=myString+systemStats()
        myString=myString+"<br><br>"
        myString=myString+"<a href='launchInfiniumCnv' target='blank'>Launch IDAT file CNV importer process in terminal window</a><br><br>"
        myString=myString+"<a href='launchInfiniumBeta' target='blank'>Launch IDAT file methylation (beta value) importer process in terminal window</a>"
        return myString

    @cherrypy.expose
    def restartServer(self):
        myString=menuheader(0,15)
        myString=myString+"NanoDiP server is restarting. It may be necesscary to reload some tabs to resume operation. Click on a menu item to proceed."
        restartNanoDiP()
        return myString
    
    @cherrypy.expose
    def resetQueue(self,queueName=""):
        myString=menuheader(0,15)
        if queueName:
            if queueName=="cpg":
                MinKnowIfPApi.cpgQueue=0
            if queueName=="umap":
                MinKnowIfPApi.umapQueue=0
            if queueName=="cnvp":
                MinKnowIfPApi.cnvpQueue=0
            if queueName=="bc":
                MinKnowIfPApi.bcQueue=0
            myString=myString+queueName+" queue reset"
        return myString
    
    @cherrypy.expose
    def listPositions(self):      
        myString=menuheader(1,10)
        positions=listMinionPositions()
        for pos in positions:
            n=str(pos.name) # pos.state does not tell much other than that the device is connected with USB ("running")
            #if n.startswith("MN"): # only list Mk1B units, not P2S
            myString=myString+"<br><iframe src='DeviceStatusLive?deviceString="+n+"' height='200' width='600' title='"+n+"' border=3></iframe>"
            myString=myString+"<iframe src='AnalysisStatusLive?deviceString="+n+"' height='200' width='600' title='"+n+"' border=3></iframe>"
            myString=myString+"<br><a href='DeviceStatusLive?deviceString="+n+"' target='_blank' title='Click to open device status page in new tab or window'>"+n+"</a>"
            myString=myString+", live state: "+getRealDeviceActivity(n)
            activeRun=getActiveRun(n)
            myString=myString+", active run: "+getActiveRun(n)
            if activeRun!="none":
                myString=myString+" <a href='launchAutoTerminator?sampleName="+getThisRunSampleID(n)+"&deviceString="+n+"&wantedBasesOverride="+str(wantedBases)+"' target='_blank'>"
                myString=myString+"<br>Click this link to launch automatic run terminator after "+str(round(wantedBases/1e6))+" MB.</a>"
                myString=myString+"<br><font color=''#ff0000'><a href='stopSequencing?deviceId="+n+"' title='Clicking this will terminate the current run immediately! Use with care!'>terminate manually</a></font>"
            myString=myString+"<br><br>"
        myString=myString+"</body></html>"
        return myString

    @cherrypy.expose
    def startSequencing(self,deviceId="",sampleId="",seqkit="",readsPerFile="",runDuration="",referenceFile="",startBiasVoltage=""):
        myString=menuheader(2,0)
        if sampleId:
            if float(runDuration)>=0.1:
                sys.argv = ['',
                            '--host','localhost',
                            '--position',deviceId,
                            '--sample-id',sampleId,
                            '--experiment-group',sampleId,
                            '--experiment-duration',runDuration,
                          #  '--basecalling',
                          #  '--fastq',
                          #  '--fastq-reads-per-file',readsPerFile,
                            '--fast5',
                            '--fast5-reads-per-file',readsPerFile,
                            '--verbose',
                            '--kit',seqkit,                     # 'SQK-RBK004', # '--kit','SQK-RBK114-24'
                          #  '--barcoding',
                          #  '--barcode-kits',seqkit,            # 'SQK-RBK004', # '--barcode-kits','SQK-RBK114-24'                            
                            '--','--start_bias_voltage',startBiasVoltage] # The "--" are required for so-called extra-arguments.
                realRunId=startRun()
                writeReferenceDefinition(sampleId,referenceFile)
                myString=myString+"sequencing run started for "+sampleId+" on "+deviceId+" as "+realRunId+" with reference "+referenceFile
                myString=myString+"<hr>"+getThisRunInformation(deviceId)
                myString=myString+'''<hr>Navigate to <b>MK1b Status</b> to launch the run terminator. It may take several minutes until the link for the
                    run terminator appears. This is due to the inexistent run state while the flow cell is being heated up to operation temperature.
                    In addition, you may want to navigate to <b>Analyze</b> and launch <b>get CpGs</b>.<br><br>
                    If you do not start the run terminator, you will have to terminate the run manually, or it will stop after the predefined time.'''
                myString=myString+'''<hr>If you see an error message above instead of a (long) list of sequencing parameters, 
                     the run was likely not started due to a mismatch of flow cell and selected sequencing kit.
                     The sequencing software always checks the present flow cell type against the sequencing kit. 
                     You can use the back button in the web browser to modify the parameters, e.g., the kit.'''
        else:    
            myString=myString+'''<form action="startSequencing" method="GET">
                Select an idle Mk1b:&nbsp;<select name="deviceId" id="deviceId">'''
            positions=listMinionPositions()
            for pos in positions:
                thisPos=pos.name
                if getRealDeviceActivity(thisPos)=="idle":
                    if getFlowCellID(thisPos)!="":
                        myString=myString+'<option value="'+thisPos+'">'+thisPos+': '+getFlowCellID(thisPos)+'</option>'
            myString=myString+'</select>'
            myString=myString+'''
                Select sequencing kit :&nbsp;<select name="seqkit" id="seqkit">'''
            for seqk in seqkits:
                myString=myString+'<option value="'+str(seqk)+'">'+str(seqk)+'</option>'
            myString=myString+'</select>'
            myString=myString+'&nbsp;<input type="text" name="readsPerFile" value="4000"/>&nbsp;'
            myString=myString+'''&nbsp; and enter the sample ID:&nbsp;<input type="text" name="sampleId" />
                &nbsp;with start voltage&nbsp;<select name="startBiasVoltage" id="startBiasVoltage">'''
            for vo in range(-180,-260,-5):
                myString=myString+'<option value="'+str(vo)+'">'+str(vo)+' mV</option>'
            myString=myString+'''</select>
                &nbsp;for&nbsp;<input type="text" name="runDuration" value="72" />&nbsp;hours.
                &nbsp;Reference set&nbsp;<select name="referenceFile" id="referenceFile">'''
            for ref in getReferenceAnnotations():
                myString=myString+'<option value="'+ref+'">'+ref+'</option>'
            myString=myString+'&nbsp;<input type="submit" value="start sequencing now"/></form>'
        return myString

    @cherrypy.expose
    def startTestrun(self,deviceId="",seqkit="",readsPerFile=""):
        myString=menuheader(3,0)
        if deviceId:
            sampleId=datetimestringnow()+"_TestRun_"+getFlowCellID(deviceId)
            sys.argv = ['',
                        '--host','localhost',
                        '--position',deviceId,
                        '--sample-id',sampleId,
                        '--experiment-group',sampleId,
                        '--experiment-duration','0.1',
                       # '--basecalling',
                       # '--fastq',
                       # '--fastq-reads-per-file',readsPerFile,
                        '--fast5',
                        '--fast5-reads-per-file',readsPerFile,
                        '--verbose',
                        '--kit',seqkit,                     # 'SQK-RBK004', # '--kit','SQK-RBK114-24'
                        #'--barcoding',
                        #'--barcode-kits',seqkit,            # 'SQK-RBK004', # '--barcode-kits','SQK-RBK114-24'
                        '--','--start_bias_voltage','-180'] # The "--" are required for so-called extra-arguments. For test runs, the default -180 mV are ok.
            realRunId=startRun()
            myString=myString+"sequencing run started for "+sampleId+" on "+deviceId+" as "+realRunId
            myString=myString+"<hr>"+getThisRunInformation(deviceId)
            myString=myString+'''<hr>If you see an error message above instead of a (long) list of sequencing parameters, 
                                 the run was likely not started due to a mismatch of flow cell and selected sequencing kit.
                                 The sequencing software always checks the present flow cell type against the sequencing kit. 
                                 You can use the back button in the web browser to modify the parameters, e.g., the kit.'''
        else:    
            myString=myString+'''<form action="startTestrun" method="GET">
                Select an idle Mk1b:&nbsp;<select name="deviceId" id="deviceId">'''
            positions=listMinionPositions()
            for pos in positions:
                thisPos=pos.name
                if getRealDeviceActivity(thisPos)=="idle":
                    if getFlowCellID(thisPos)!="":
                        myString=myString+'<option value="'+thisPos+'">'+thisPos+': '+getFlowCellID(thisPos)+'</option>'
            myString=myString+'</select>'
            myString=myString+'''
                Select sequencing kit :&nbsp;<select name="seqkit" id="seqkit">'''
            for seqk in seqkits:
                myString=myString+'<option value="'+str(seqk)+'">'+str(seqk)+'</option>'
            myString=myString+'</select>'
            myString=myString+'&nbsp;<input type="text" name="readsPerFile" value="4000"/>&nbsp;'
            myString=myString+'''
                </select>&nbsp;<input type="submit" value="start test run now (0.1h)"/></form>'''
        return myString
    
    @cherrypy.expose
    def stopSequencing(self,deviceId=""):      
        myString=menuheader(1,0)
        myString=myString + stopRun(deviceId)
        myString=myString + "<br><br>Click on any menu item to proceed."
        return myString
    
    @cherrypy.expose
    def changeVoltageLive(self,deviceId="",newVoltage=""):      
        myString=menuheader(1,0)
        myString=myString + setBiasVoltage(deviceId,newVoltage)
        myString=myString + "<br><br>Click on any menu item to proceed."
        return myString
    
    @cherrypy.expose
    def listExperiments(self):
        myString=menuheader(4,10)
        myString=myString+"Running and buffered experiments:<br>"
        experiments=listMinionExperiments()
        myString=myString+experiments
        return myString 
    
    @cherrypy.expose
    def listRuns(self):
        myString=menuheader(5,0)
        myString=myString+makePastAnalysesTable()
        return myString    
      
    @cherrypy.expose
    def analyze(self):
        myString=menuheader(6,0)        
        myString=myString+analysisLaunchTable()
        return myString
 
    @cherrypy.expose
    def cnvplot(self, sampleName=None):
        myString=""
        if sampleName:
            while MinKnowIfPApi.cnvpQueue>maxcnvqueue:
                time.sleep(2)
            MinKnowIfPApi.cnvpQueue+=1
            try:
                createCNVPlot(sampleName)
                errorString=""
            except:
                errorString="<b><font color='#FF0000'>ERROR OCCURRED, PLEASE RELOAD TAB</font></b>"
            myString="<html><head><title>"+sampleName+" at "+datetimestringnow()+"</title></head><body>"
            myString=myString+errorString
            myString=myString+"<img src='reports/"+sampleName+"_CNVplot.png' width='100%'>"
            myString=myString+"</body></html>"
            MinKnowIfPApi.cnvpQueue-=1
        return myString
    
    @cherrypy.expose
    def umapplot(self, sampleName=None, refAnno=None):
        myString=""
        if sampleName and refAnno:
            while MinKnowIfPApi.umapQueue>0:
                time.sleep(2)
            MinKnowIfPApi.umapQueue+=1
            refAnnoName=refAnno.replace(".xlsx","")
            try:
                methylationUMAP(sampleName,refAnno)
                errorString=""
            except:
                errorString="<b><font color='#FF0000'>ERROR OCCURRED, PLEASE RELOAD TAB</font></b>"
            myString="<html><head><title>"+sampleName+" against "+refAnno+" at "+datetimestringnow()+"</title>"
            myString=myString+"<meta http-equiv='refresh' content='1; URL=reports/"+sampleName+"_"+refAnnoName+"_UMAP_all.html'>"
            myString=myString+"</head><body>"
            myString=myString+errorString
            myString=myString+"Loading UMAP plot. It it fails, <a href='reports/"+sampleName+"_"+refAnnoName+"_UMAP_all.html'>click here to load plot</a>."
            myString=myString+"</body></html>"
            MinKnowIfPApi.umapQueue-=1
        return myString
    
    @cherrypy.expose
    def makePdf(self, sampleName=None, refAnno=None):    
        if sampleName and refAnno:
            myString=generatePdfReport(sampleName,refAnno,"NanoDiP")
        return myString
    
    @cherrypy.expose
    def epidipGenUmap(self, stdFile=None, referenceFile=None, topN=None):
        myString=menuheader(7,0)
        if stdFile and referenceFile and topN:
            myString=myString+"EpiDiP UMAP, stdDev="+stdFile+", reference="+referenceFile+" topN probes="+topN+"<hr>"
            calculateStdev(referenceFile)
            e=epidipUmap(referenceFile, stdFile, int(topN))
            plotUmap(e)
            myString=myString+"Output saved in "+e
        else:
            myString=myString+'<form action="epidipGenUmap" method="GET">'
            myString=myString+'</select>Reference set&nbsp;<select name="referenceFile" id="referenceFile">'
            for ref in getReferenceAnnotations():
                myString=myString+'<option value="'+ref+'">'+ref+'</option>'
            myString=myString+'</select>&nbsp;StdDev file&nbsp;<select name="stdFile" id="stdFile">'
            for ref in getReferenceAnnotations():
                myString=myString+'<option value="'+ref+'">'+ref+'</option>'
            myString=myString+'</select>'
            myString=myString+'&nbsp;top n probes:&nbsp;<input type="text" name="topN" id="topN" value="5000"/>&nbsp;'
            myString=myString+'&nbsp;<input type="submit" value="re-generate UMAP"/></form>'
            myString=myString+'<hr>Note that UMAP re-generation can take a while.'
        myString=myString+"<hr>Past results<br><br>"+makePastUmapTable()
        return myString    

    @cherrypy.expose
    def epidipAnnoUmap(self, epidipUmapXlsx=None, sentrixList=None):
        myString=menuheader(8,0)
        myString=myString+epidipUmapAnnoPage(epidipUmapXlsx,sentrixList)
        return myString
    
    @cherrypy.expose
    def epidipCnv(self, sentrixId=None, searchGenes=None):
        myString=menuheader(9,0)
        myString=myString+cnvPlotPage(sentrixId, searchGenes)
        return myString
        
    @cherrypy.expose
    def epidipReport(self, sentrixid=None, referenceFile=None, umapFile=None):
        myString=menuheader(10,0)
        if sentrixid and referenceFile and umapFile:
            myString=myString+"EpiDiP report for "+sentrixid+" with "+referenceFile+":<hr>"
            myString=myString+generateEpidipReport(sentrixid, referenceFile, umapFile)
        else:
            myString=myString+'<form action="epidipReport" method="GET">'
            myString=myString+'SentrixID:&nbsp;<input type="text" name="sentrixid" />&nbsp;'
            myString=myString+'UMAP coordinates&nbsp;<select name="umapFile" id="umapFile">'
            for u in epidipUmapCoordUrlFiles:
                myString=myString+'<option value="'+u+'">'+u+'</option>'
            myString=myString+'</select>&nbsp;Reference set&nbsp;<select name="referenceFile" id="referenceFile">'
            for ref in getReferenceAnnotations():
                myString=myString+'<option value="'+ref+'">'+ref+'</option>'
            myString=myString+'&nbsp;<input type="submit" value="generate report"/></form>'
            myString=myString+'<hr>Note that report generation can take a while. A web access log will be displayed upon completion, documenting the communication with the EpiDiP server.'
        return myString 
    
    
    @cherrypy.expose
    def xdipReport(self, sentrixid=None, referenceFile=None, umapFile=None):
        myString=menuheader(11,0)
        if sentrixid and refAnno and umapCoordFile:
            myString=myString+"xDiP report for <b>"+sentrixid+"</b> with <b>"+refAnno+"</b> annotation based on <b>" +umapCoordFile+ "</b> coordinates."
        else:
            myString=myString+"Generate xDiP report."
            myString=myString+'<form action="epidipReport" method="GET">'
            myString=myString+'SentrixID:&nbsp;<input type="text" name="sentrixid" /><br>'
            myString=myString+'Reference Annotation:&nbsp;<select name="referenceFile" id="referenceFile">'
            refAnnos=[]
            for dn,snd,filenames in os.walk(referenceDir):
                for fn in filenames:
                    if fn.endswith(".xlsx"):
                        refAnnos.append(fn)
            refAnnos.sort()
            for rA in refAnnos:            
                myString=myString+'<option value="'+rA+'">'+rA+'</option>'
            myString=myString+'</select><br>'
            myString=myString+'UMAP coordinates&nbsp;<select name="umapFile" id="umapFile">'
            umapFiles=[]
            for dn,snd,filenames in os.walk(epidipTmp):
                for fn in filenames:
                    if fn.endswith(".xlsx"):
                        umapFiles.append(fn)
            umapFiles.sort(reverse=True)
            for uF in umapFiles:            
                myString=myString+'<option value="'+uF+'">'+uF+'</option>'            
            myString=myString+'</select><br>'
            myString=myString+'&nbsp;<input type="submit" value="generate report"/></form>'
        return myString
    
    @cherrypy.expose
    def cleanup(self,deltype="",deldir=""):
        myString=menuheader(12,0)+cleanuptable(deltype,deldir)
        return myString
    
    @cherrypy.expose
    def about(self):
        myString=menuheader(13,0)+'''
        <b>NanoDiP</b> is a tool to obtain and analyze low-coverage whole genome
        nanopore sequencing information through bascalling, genomic alignment,
        copy number extrapolation, and unsupervised machine learning by UMAP-based
        dimensions reduction. It is the nanopore-centered implementation of
        <a href="http://www.epidip.org">EpiDiP</a> which stands for <i>Epigenomic
        Digital Pathology</i>. NanoDiP hence abbreviates <i>Nanopore
        Digital Pathology</i>.<br><br>
        Nanopore sequencing is developed and sold by <a href="https://nanoporetech.com/">ONT</a>.
        Then authors of this software are not affiliated with ONT.<br><br>
        This software is licensed under the 
        <a href="https://www.gnu.org/licenses/gpl-3.0.html">GPLv3</a>. By using this
        program, you agree to the terms specified herein.<br><br>
        <b>This software is not a medical device.</b> Its use occurs in the sole responsibility
        of the treating physician. The authors shall not be held liable for any
        damage caused by this software. 
        <b>Basic understanding of how this system works and internal validation are 
        strongly advised before implementation in a diagnostic setting.</b>
        '''
        return myString  
    
    @cherrypy.expose
    def DeviceStatusLive(self,deviceString=""):
        currentFlowCellId=getFlowCellID(deviceString)
        myString="<html><head><title>"+deviceString+": "+currentFlowCellId+"</title>"
        try:
            myString=myString+"<meta http-equiv='refresh' content='2'>"
            if getRealDeviceActivity(deviceString)=="sequencing":
                myString=myString+"<body bgcolor='#00FF00'>"
            else:
                myString=myString+"<body>"
            myString=myString+"<b>"+deviceString+": "+currentFlowCellId+"</b><br><tt>"
            myString=myString+getMinKnowApiStatus(deviceString)
        except:
            myString=myString+"<br>No previous device activity, information will appear as soon as the device has been running once in this session.<br>"
        myString=myString+"Sample ID: "+getThisRunSampleID(deviceString)+"<br>"
        myString=myString+getThisRunState(deviceString)
        myString=myString+"<br>"+getThisRunYield(deviceString)
        myString=myString+"</tt></body></html>"
        return myString
    
    @cherrypy.expose
    def AnalysisStatusLive(self,deviceString=""):
        myString=""
        if deviceString:
            myString=livePage(deviceString)
        return myString

    @cherrypy.expose
    def analysisLauncher(self,functionName="",sampleName="",refAnno="",basecallModel="",barcodeSet=""):
        if functionName and sampleName and refAnno:
            myString="<html><head><title>"+sampleName+" "+functionName+"</title></head><body>"
            myString=myString+functionName+" launched for "+sampleName+" "
            if refAnno!="None":
                myString=myString+"against "+refAnno
            myString=myString+" at "+datetimestringnow()+". "
            myString=myString+"Frame below will display result upon completion, if this tab/window is kept open."
            if refAnno=="None":
                myString=myString+"<br><iframe src='./"+functionName+"?sampleName="+sampleName+"' height='95%' width='100%' title='"+sampleName+"' border=3></iframe>"
            elif refAnno=="includeunclassified" or refAnno=="predominantbarcode":
                myString=myString+"<br><iframe src='./"+functionName+"?sampleName="+sampleName+"&refAnno="+refAnno+"' height='95%' width='100%' title='"+sampleName+"' border=3></iframe>"
            else:
                myString=myString+"<br><iframe src='./"+functionName+"?sampleName="+sampleName+"&refAnno="+refAnno+"' height='95%' width='100%' title='"+sampleName+"' border=3></iframe>"
        elif functionName and sampleName and basecallModel and barcodeSet:
            myString="<html><head><title>"+sampleName+" "+functionName+"</title></head><body>"
            myString=myString+functionName+" launched for "+sampleName+" "
            myString=myString+" at "+datetimestringnow()+". "
            myString=myString+"Frame below will display result upon completion, if this tab/window is kept open."
            myString=myString+"<br><iframe src='./"+functionName+"?sampleName="+sampleName+"&basecallModel="+basecallModel+"&barcodeSet="+barcodeSet+"' height='95%' width='100%' title='"+sampleName+"' border=3></iframe>"
        else:
            myString="Nothing to launch. You may close this tab now."
        return myString
    
    @cherrypy.expose
    def analysisPoller(self,sampleName="",deviceString="",runId=""):
        myString="<html><head>"
        if sampleName and deviceString and runId:
                myString=myString+"<title>Poller: "+sampleName+"/"+deviceString+"/"+runId+"</title>"
                myString=myString+"<meta http-equiv='refresh' content='15'>"
                myString=myString+"<body>"
                myString=myString+"Last refresh for "+sampleName+"/"+deviceString+"/"+runId+" at "+datetimestringnow()
                myString=myString+"</body></html>"
                writeRunTmpFile(sampleName,deviceString)
        return myString

    @cherrypy.expose
    def basecallPoller(self,sampleName="",basecallModel="",barcodeSet=""):
        while MinKnowIfPApi.bcQueue>0:
            time.sleep(2)
        MinKnowIfPApi.bcQueue+=1
        myString=basecallLivePage(sampleName,basecallModel,barcodeSet)
        MinKnowIfPApi.bcQueue-=1
        return myString
    
    @cherrypy.expose
    def methylationPoller(self,sampleName="",refAnno=""):
        while MinKnowIfPApi.cpgQueue>maxcpgqueue:
            time.sleep(2)
        MinKnowIfPApi.cpgQueue+=1
        myString=methcallLivePage(sampleName,refAnno)
        MinKnowIfPApi.cpgQueue-=1
        return myString

    @cherrypy.expose
    def launchAutoTerminator(self,sampleName="",deviceString="",wantedBasesOverride=str(wantedBases)):
        myString="ERROR"
        if sampleName and deviceString:
            myString=thisRunWatcherTerminator(deviceString,sampleName,wantedBasesOverride)
        return myString

    @cherrypy.expose
    def launchInfiniumCnv(self):
        myString=inifiniumCnvLauncher()
        return myString
    
    @cherrypy.expose
    def launchInfiniumBeta(self):
        myString=inifiniumBetaLauncher()
        return myString

    @cherrypy.expose
    def shutdown(self):  
        cherrypy.engine.exit()
           
# CherryPy server configuration
config = {  
    'global' : {
        'server.socket_host' : cherrypyHost,
        'server.socket_port' : cherrypyPort,
        'server.thread_pool' : cherrypyThreads,
        'response.timeout' : 60,
        'server.shutdown_timeout': 1 
              }
}
# Start CherryPy Webserver
if debugLogging==True:
    cherrypy.log.screen=True #set access logging
    cherrypy.config.update({'log.screen': True})
else:
    cherrypy.log.screen=False #set access logging
    cherrypy.config.update({'log.screen': False})
    cherrypy.config.update({ "environment": "embedded" })
cherrypy.config.update(config)
print("NanoDiP server running at http://"+cherrypyHost+":"+str(cherrypyPort))
print("GPU present: "+str(gpu))
if __name__ == '__main__':
#    cherrypy.tree.mount(root, '/', config=config) 
    cherrypy.quickstart(MinKnowIfPApi(),
                        '/',
                        {'/favicon.ico':
                         {'tools.staticfile.on':True,
                          'tools.staticfile.filename':
                          thisFaviconPath},
                        '/img':
                         {'tools.staticdir.on': True,
                          'tools.staticdir.dir': 
                          imgPath},
                        '/reports':
                         {'tools.staticdir.on': True,
                          'tools.staticdir.dir': 
                          nanodipReportDir},
                        '/cnv':
                         {'tools.staticdir.on': True,
                          'tools.staticdir.dir': 
                          cnvHtmlPath},
                        '/epidip':
                         {'tools.staticdir.on': True,
                          'tools.staticdir.dir': 
                          epidipTmp}})


# ### ^^^ LIVE LOG ABOVE ^^^ 
# All CherryPy access will be logged here, including live progress bars for computationally intense analyses. Detailed access logging is turned off by default (accessLogging is False), but can be turned on,e.g., for debugging, in the configuration section at the beginning of this notebook. While it is not required to have at look at these during normal operation, information contained in the log may be helpful in troubleshooting. Line numbers in error messages indicated here typically match those given in the respective Jupyter Notebook cells.
# 
# To preseve these messages, halt the Python kernel, save and close the notebook to send it for support. This makes sure that the code as well as the error messages will be preserved.
# 
# To launch the user interface, wait until you see a pink log entry that the web server has started, then navigate to http://localhost:8081.

# # CONSTRUCTION

# In[88]:


#


# # ToDo
# * prevent duplicate execution DONE
# * prevent repeated execution DONE
# * count repeated failed execution of single ONTfile processing and give up after n trials DONE
# * n trials to config DONE
# * tell "done" if all present ONTfiles have been processed. DONE
# * report indicate predominant barcode DONE
# * GPU-accel. mm2 DONE
# * set MBases for launch by argument (GUI not required or in run terminator page) DONE
# * handle POD5 properly (conversion for f5c)
# * handle rapid non-barcoded kits - potentially by adding them to the list of possible kits

# In[ ]:





# In[ ]:




